# takeoffnsw_formatter.py
import streamlit as st
import pandas as pd
import io
import re
import json
import csv
from datetime import datetime
from zipfile import ZipFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------
# Utility functions
# ---------------------------

def detect_columns(df):
    lookup = {c.strip().upper(): c for c in df.columns}
    def pick(*names):
        for n in names:
            if n.upper() in lookup:
                return lookup[n.upper()]
        return None
    return {
        'PRODUCT': pick('Subject','PRODUCT','Item','ITEM','NAME'),
        'SCHEDULE BRAND': pick('Manufacturer','BRAND','MAKE','SCHEDULE BRAND'),
        'SCHEDULE MODEL': pick('Model','MODEL','MODEL #','SCHEDULE MODEL'),
        'BRAND': pick('Manufacturer','BRAND','MAKE'),
        'MODEL': pick('Model','MODEL','MODEL #'),
        'QTY': pick('Count','QTY','Quantity','QUANT'),
        'TAG': pick('Label','TAG','Mark'),
        'NECK SIZE': pick('Neck Size','NECK SIZE','NECKSIZE'),
        'MODULE SIZE': pick('Face Size','FACE','MODULE SIZE'),
        'DUCT SIZE': pick('Duct Size','DUCTSIZE'),
        'TYPE': pick('Type','DESCRIPTION','TYPE'),
        'MOUNTING': pick('Mounting','MOUNT'),
        'ACCESSORIES1': pick('Accessories','ACCESSORIES','ACCESSORIES1'),
        'ACCESSORIES2': pick('Description','DESCRIPTION'),
        'REMARK': pick('Remark','Remarks','Notes','REMARK')
    }

def pad_numeric_only(tag):
    s = str(tag).strip()
    return s.zfill(2) if s.isdigit() else s

_token_re = re.compile(r'(\d+|\D+)')
def tag_sort_key(tag):
    t = str(tag).strip()
    if t == '':
        return (1, '', ())
    if re.match(r'^[A-Za-z]', t):
        parts = _token_re.findall(t)
        seq = tuple((0, p.lower()) if not p.isdigit() else (1, int(p)) for p in parts)
        return (0, t.lower(), seq)
    m = re.match(r'^(\d+)', t)
    if m:
        num = int(m.group(1))
        rest = t[m.end():].lower()
        return (2, num, rest)
    return (1, t.lower(), ())

def neck_size_key(val):
    m = re.search(r'(\d+(\.\d+)?)', str(val))
    return float(m.group(1)) if m else float('inf')

def mounting_rank(val):
    if not isinstance(val, str):
        return 2
    v = val.lower()
    lay_variants = ['lay-in','lay in','in lay','in laying','in-lay','layin','laying','inlay','lay']
    for variant in lay_variants:
        if variant in v:
            return 0
    if 'surface' in v:
        return 1
    return 2

def to_float_safe(x):
    try:
        s = str(x).strip()
        return float(s) if s!='' else 0.0
    except:
        return 0.0

# Robust upload reader
def read_uploaded_file(uploaded_file):
    """
    Robust reader for uploaded files (Streamlit).
    Returns (df, error_message). If df is None, error_message contains details.
    """
    if uploaded_file is None:
        return None, "No file uploaded."

    name = uploaded_file.name.lower()
    # Excel first
    if name.endswith(('.xls', '.xlsx')):
        try:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, dtype=str)
            if df.shape[1] == 0:
                return None, "Excel file read OK but no columns were detected."
            df = df.fillna('')
            return df, None
        except Exception as e:
            return None, f"Failed to read Excel file: {e}"

    # Otherwise treat as CSV-like
    try:
        uploaded_file.seek(0)
        raw = uploaded_file.read()
        if isinstance(raw, bytes):
            # try utf-8-sig then latin1
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                try:
                    text = raw.decode('latin1')
                except Exception:
                    return None, "Failed to decode CSV file bytes as UTF-8 or Latin-1."
        else:
            text = str(raw)

        # remove blank lines
        lines = [ln for ln in text.splitlines() if ln.strip() != '']
        if len(lines) == 0:
            return None, "CSV appears to be empty (no non-blank lines)."

        sample = "\n".join(lines[:10])

        # Try csv.Sniffer to detect delimiter
        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            delim = dialect.delimiter
        except Exception:
            for d in [',',';','\t','|']:
                if d in sample:
                    delim = d
                    break
            else:
                delim = ','

        # Read with pandas using detected delimiter; fallback to sep=None if needed
        try:
            df = pd.read_csv(io.StringIO(text), sep=delim, dtype=str, engine='python')
        except Exception as e:
            try:
                df = pd.read_csv(io.StringIO(text), sep=None, engine='python', dtype=str)
            except Exception as e2:
                return None, f"Failed to parse CSV: {e} (fallback error: {e2})"

        if df.shape[1] == 0:
            return None, "CSV read OK but no columns were detected."
        df = df.fillna('')
        return df, None

    except Exception as e:
        return None, f"Unexpected failure reading file: {e}"

# Styling & Excel save
def style_and_save_excel(df, excel_bytes_io,
                         header_fill="#CFE2F3", subtotal_fill="#FFFF00", grand_fill="#CFE2F3",
                         bold_cols = ['PRODUCT','SCHEDULE MODEL','TAG','MODULE SIZE','TYPE','ACCESSORIES1']):
    wb = Workbook()
    ws = wb.active
    ws.title = "TakeoffNSW"

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = PatternFill(start_color=header_fill.replace('#',''), end_color=header_fill.replace('#',''), fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            colname = df.columns[c_idx-1]
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if colname in bold_cols and str(value).strip() != '':
                cell.font = Font(bold=True)
            if colname == 'QTY' and str(value).strip() != '':
                try:
                    tag_val = str(row[df.columns.get_loc('TAG')]).strip() if 'TAG' in df.columns else ''
                except Exception:
                    tag_val = ''
                prod_val = str(row[0]).strip()
                if tag_val.endswith('(TOTAL)') or prod_val == 'Grand Total' or ('=' in prod_val and prod_val.split('=')[-1].strip().replace('.','',1).isdigit()):
                    cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=False)
            if colname == 'TAG' and isinstance(value, str) and value.endswith('(TOTAL)'):
                for name in ['PRODUCT','TAG','QTY']:
                    i = list(df.columns).index(name) + 1
                    ccell = ws.cell(row=r_idx, column=i)
                    ccell.fill = PatternFill(start_color=subtotal_fill.replace('#',''), end_color=subtotal_fill.replace('#',''), fill_type="solid")
                    ccell.font = Font(bold=True)
                    ccell.border = border
            if value == 'Grand Total':
                for col_i in range(1, len(df.columns)+1):
                    ccell = ws.cell(row=r_idx, column=col_i)
                    ccell.fill = PatternFill(start_color=grand_fill.replace('#',''), end_color=grand_fill.replace('#',''), fill_type="solid")
                    ccell.font = Font(bold=True)
                    ccell.border = border

    ws.freeze_panes = 'A2'
    for i, col in enumerate(df.columns, 1):
        try:
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        except Exception:
            max_len = len(col) + 2
        ws.column_dimensions[get_column_letter(i)].width = max_len

    wb.save(excel_bytes_io)
    excel_bytes_io.seek(0)
    return excel_bytes_io

# ---------------------------
# Core takeoff builder
# ---------------------------

def build_takeoff(df_in,
                  mapping,
                  apply_brand_autofill=True,
                  keep_model_blank=True,
                  replace_empty_with_dot=False,
                  aggregate_duplicates=True):
    target_cols = ['PRODUCT','SCHEDULE BRAND','SCHEDULE MODEL','BRAND','MODEL','QTY','TAG',
                   'NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2','REMARK']

    # Map columns
    df = pd.DataFrame()
    for c in target_cols:
        src = mapping.get(c)
        if src and src in df_in.columns:
            df[c] = df_in[src].astype(str).fillna('')
        else:
            df[c] = ''

    # Brand autofill (only those specific rules)
    if apply_brand_autofill:
        prod_upper = df['PRODUCT'].astype(str).str.upper()
        df['BRAND'] = ''
        df.loc[prod_upper.str.contains('AD-GRD', na=False), 'BRAND'] = 'PRICE'
        df.loc[prod_upper.str.contains('FAN', na=False), 'BRAND'] = 'LOREN COOK'
        df.loc[prod_upper.str.contains('SPLIT SYSTEM', na=False), 'BRAND'] = 'SAMSUNG'
    else:
        if mapping.get('BRAND') and mapping['BRAND'] in df_in.columns:
            df['BRAND'] = df_in[mapping['BRAND']].astype(str).fillna('')
        else:
            df['BRAND'] = ''

    # MODEL blank handling
    if keep_model_blank:
        df['MODEL'] = ''
    else:
        if mapping.get('MODEL') and mapping['MODEL'] in df_in.columns:
            df['MODEL'] = df_in[mapping['MODEL']].astype(str).fillna('')

    # Replace empties with dot if requested
    if replace_empty_with_dot:
        df = df.replace('', '.').replace('nan','.')

    # numeric qty and tag normalization
    df['QTY'] = df['QTY'].apply(to_float_safe)
    df['TAG'] = df['TAG'].astype(str).apply(lambda t: pad_numeric_only(t.strip()))

    # optional aggregation of exact duplicate attribute rows
    if aggregate_duplicates:
        agg_keys = ['PRODUCT','TAG','SCHEDULE BRAND','SCHEDULE MODEL','NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2']
        for k in agg_keys:
            if k not in df.columns:
                df[k] = ''
        grouped = df.groupby(agg_keys, dropna=False, as_index=False).agg({
            'BRAND': 'first',
            'MODEL': 'first',
            'QTY': 'sum',
            'REMARK': lambda x: '; '.join([v for v in x.astype(str).unique() if v and v!='nan'])
        })
        new_df = pd.DataFrame()
        for c in target_cols:
            new_df[c] = grouped.get(c, '')
        df = new_df

    # Sorting helpers and sort
    df['_prod_key'] = df['PRODUCT'].fillna('').astype(str).str.lower()
    df['_tag_key'] = df['TAG'].apply(tag_sort_key)
    df['_neck'] = df['NECK SIZE'].apply(neck_size_key)
    df['_mount'] = df['MOUNTING'].apply(mounting_rank)

    df = df.sort_values(by=['_prod_key','_tag_key','_neck','_mount'], kind='mergesort').reset_index(drop=True)

    # Hybrid grouping: AD-GRD by TAG; others by PRODUCT
    out_rows = []
    grand_total = 0.0
    input_tags = set(df['TAG'].unique())

    for prod_key, group in df.groupby('_prod_key', sort=True):
        group = group.copy().reset_index(drop=True)
        if group.empty:
            continue
        prod_name = group.loc[0,'PRODUCT']
        is_adgrd = isinstance(prod_name, str) and 'AD-GRD' in prod_name.upper()
        if is_adgrd:
            first_of_product = True
            tags = group['TAG'].unique().tolist()
            for tag in tags:
                tag_group = group[group['TAG'] == tag]
                subtotal = tag_group['QTY'].sum()
                grand_total += subtotal
                for idx, r in tag_group.iterrows():
                    row = {c: r[c] if c in r.index else '' for c in target_cols}
                    row['PRODUCT'] = prod_name if first_of_product else ''
                    out_rows.append(row)
                    first_of_product = False
                subtotal_row = {c: '' for c in target_cols}
                subtotal_row['PRODUCT'] = f"{tag}={int(subtotal) if subtotal == int(subtotal) else '{:.2f}'.format(subtotal)}"
                subtotal_row['TAG'] = f"{tag} (TOTAL)"
                subtotal_row['QTY'] = subtotal
                out_rows.append(subtotal_row)
            out_rows.append({c: '' for c in target_cols})
        else:
            first_of_product = True
            prod_total = 0.0
            for idx, r in group.iterrows():
                row = {c: r[c] if c in r.index else '' for c in target_cols}
                row['PRODUCT'] = prod_name if first_of_product else ''
                out_rows.append(row)
                first_of_product = False
                prod_total += r['QTY']
                grand_total += r['QTY']
            subtotal_row = {c: '' for c in target_cols}
            subtotal_row['PRODUCT'] = f"{prod_name}={int(prod_total) if prod_total == int(prod_total) else '{:.2f}'.format(prod_total)}"
            subtotal_row['TAG'] = f"{prod_name.split('=')[0] if '=' in prod_name else prod_name.split('=')[0]} (TOTAL)"
            subtotal_row['QTY'] = prod_total
            out_rows.append(subtotal_row)
            out_rows.append({c: '' for c in target_cols})

    # grand total row
    grand_row = {c: '' for c in target_cols}
    grand_row['PRODUCT'] = 'Grand Total'
    grand_row['TAG'] = 'Grand Total'
    grand_row['QTY'] = grand_total
    out_rows.append(grand_row)

    final_df = pd.DataFrame(out_rows, columns=target_cols)

    # integrity checks
    output_tags = set([t for t in final_df['TAG'].dropna() if isinstance(t,str) and t.strip()!='' and not t.endswith('(TOTAL)')])
    missing_tags = sorted(list(input_tags - output_tags))
    integrity = {
        "input_unique_tags": len(input_tags),
        "output_unique_tags": len(output_tags) if 'output_tags' in locals() else len(output_tags) if 'output_tags' in globals() else len(output_tags) if False else len(output_tags) if False else len(output_tags) if False else len(output_tags) if False else len(output_tags)  # placeholder (will be replaced below)
    }
    # (safer to recompute properly)
    integrity = {
        "input_unique_tags": len(input_tags),
        "output_unique_tags": len(output_tags),
        "all_tags_preserved": len(missing_tags) == 0,
        "missing_tags_sample": missing_tags[:20]
    }

    return final_df, integrity

# ---------------------------
# Streamlit UI
# ---------------------------

st.set_page_config(page_title="TakeoffNSW Formatter", layout="wide")
st.title("TakeoffNSW Formatter — Final")

st.markdown("Upload a raw CSV/XLSX takeoff and convert to the formatted TakeoffNSW file.")

uploaded = st.file_uploader("Upload CSV or Excel", type=['csv','xlsx','xls'])

st.sidebar.header("Options")
apply_brand = st.sidebar.checkbox("Apply brand autofill rules (AD-GRD/FAN/SPLIT SYSTEM)", value=True)
keep_model_blank = st.sidebar.checkbox("Keep MODEL blank", value=True)
replace_empty = st.sidebar.checkbox("Replace empty cells with '.'", value=False)
aggregate_duplicates = st.sidebar.checkbox("Aggregate exact duplicate attribute rows (sum QTY)", value=True)

st.sidebar.markdown("### Colors")
header_color = st.sidebar.color_picker("Header fill", "#CFE2F3")
subtotal_color = st.sidebar.color_picker("Subtotal fill", "#FFFF00")
grand_color = st.sidebar.color_picker("Grand total fill", "#CFE2F3")

st.sidebar.markdown("### Bold columns (comma-separated)")
bold_input = st.sidebar.text_input("Columns to bold (exact names)", value="PRODUCT,SCHEDULE MODEL,TAG,MODULE SIZE,TYPE,ACCESSORIES1")
bold_cols = [c.strip() for c in bold_input.split(',') if c.strip()]

st.sidebar.markdown("### Mapping editor")
if 'mapping_preset' not in st.session_state:
    st.session_state['mapping_preset'] = {}

mapping_json = st.sidebar.text_area("Mapping JSON (target -> source). Leave empty to auto-detect.", value="", height=140)
if st.sidebar.button("Load mapping from JSON"):
    try:
        m = json.loads(mapping_json)
        st.session_state['mapping_preset'] = m
        st.sidebar.success("Mapping loaded into session.")
    except Exception as e:
        st.sidebar.error(f"Invalid JSON: {e}")

if st.sidebar.button("Clear mapping saved in session"):
    st.session_state['mapping_preset'] = {}
    st.sidebar.info("Mapping reset.")

st.sidebar.markdown("### Unit split (optional)")
unit_col_guess = None
peek_df = None

# Read uploaded file early with robust function
df_in = None
if uploaded:
    df_in, read_err = read_uploaded_file(uploaded)
    if read_err:
        st.error(f"Failed to read file: {read_err}")
        st.stop()
    else:
        # prepare a peek for unit column choices
        try:
            peek_df = df_in.head(3)
        except Exception:
            peek_df = None

unit_split = st.sidebar.checkbox("Enable unit split (generate one file per unit)", value=False)
unit_col = st.sidebar.selectbox("Unit column (if detected)", options=[None] + (list(peek_df.columns) if peek_df is not None else []), index=0)
include_empty_units = st.sidebar.checkbox("Include rows with empty units as their own file", value=False)

if uploaded and df_in is not None:
    auto_map = detect_columns(df_in)
    mapping_use = auto_map.copy()
    if st.session_state.get('mapping_preset'):
        for k,v in st.session_state['mapping_preset'].items():
            mapping_use[k] = v

    st.subheader("Column mapping (target → source)")
    with st.form("mapping_form"):
        cols = ['PRODUCT','SCHEDULE BRAND','SCHEDULE MODEL','BRAND','MODEL','QTY','TAG','NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2','REMARK']
        user_map = {}
        for c in cols:
            choices = [''] + list(df_in.columns)
            default = mapping_use.get(c) if mapping_use.get(c) in df_in.columns else ''
            user_map[c] = st.selectbox(c, choices, index=(choices.index(default) if default in choices else 0), key=f"map_{c}")
        save_map = st.form_submit_button("Apply mapping")

    if save_map:
        mapping_use = user_map
        st.success("Mapping applied.")

    preview_build = pd.DataFrame()
    for c in ['PRODUCT','SCHEDULE BRAND','SCHEDULE MODEL','BRAND','MODEL','QTY','TAG','NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2','REMARK']:
        src = mapping_use.get(c)
        if src and src in df_in.columns:
            preview_build[c] = df_in[src].astype(str).fillna('')
        else:
            preview_build[c] = ''

    st.markdown("**Preview of mapped data (first 20 rows)**")
    st.dataframe(preview_build.head(20), height=300)

    if st.button("Generate Takeoff files (preview + download)"):
        with st.spinner("Building takeoff..."):
            final_df, integrity = build_takeoff(
                df_in,
                mapping_use,
                apply_brand_autofill=apply_brand,
                keep_model_blank=keep_model_blank,
                replace_empty_with_dot=replace_empty,
                aggregate_duplicates=aggregate_duplicates
            )

            st.subheader("Integrity check")
            st.json(integrity)

            st.subheader("Final Takeoff preview (first 200 rows)")
            st.dataframe(final_df.head(200), height=400)

            excel_io = io.BytesIO()
            style_and_save_excel(final_df, excel_io,
                                 header_fill=header_color, subtotal_fill=subtotal_color, grand_fill=grand_color,
                                 bold_cols=bold_cols)

            csv_io = io.BytesIO()
            csv_io.write(final_df.to_csv(index=False).encode('utf-8'))
            csv_io.seek(0)

            if unit_split and unit_col and unit_col in df_in.columns:
                units = df_in[unit_col].fillna('').unique().tolist()
                if not include_empty_units:
                    units = [u for u in units if str(u).strip()!='']
                zip_bytes = io.BytesIO()
                with ZipFile(zip_bytes, 'w') as zf:
                    for u in units:
                        mask = (df_in[unit_col].fillna('') == u)
                        sub_df = df_in[mask]
                        if sub_df.empty and not include_empty_units:
                            continue
                        sub_final, _ = build_takeoff(sub_df, mapping_use, apply_brand_autofill=apply_brand, keep_model_blank=keep_model_blank, replace_empty_with_dot=replace_empty, aggregate_duplicates=aggregate_duplicates)
                        sub_excel_io = io.BytesIO()
                        style_and_save_excel(sub_final, sub_excel_io, header_fill=header_color, subtotal_fill=subtotal_color, grand_fill=grand_color, bold_cols=bold_cols)
                        name = f"TakeoffNSW_{str(u).strip() or 'EMPTY'}.xlsx"
                        zf.writestr(name, sub_excel_io.getvalue())
                zip_bytes.seek(0)
                st.download_button("Download ZIP (one file per unit)", data=zip_bytes, file_name=f"TakeoffNSW_units_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")
            else:
                st.download_button("Download styled Excel", data=excel_io, file_name=f"TakeoffNSW_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                st.download_button("Download CSV", data=csv_io, file_name=f"TakeoffNSW_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

            st.success("Files generated. Download above.")


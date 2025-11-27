# takeoffnsw_with_units_manual_ordering.py
"""
Single-file Streamlit app:
- TakeoffNSW Formatter (Triune-compatible output)
- Data Unit Sheet builder (Triune helpers)
- Manual PRODUCT ordering (textarea + Apply) and Per-product TAG ordering (textarea + Apply)
Save and run with:
    streamlit run takeoffnsw_with_units_manual_ordering.py
"""

import io
import re
import json
import csv
import zipfile
from datetime import datetime
from typing import Optional, Tuple, Dict, List

import pandas as pd
import numpy as np
import streamlit as st

# optional st_aggrid (kept safe, not required)
try:
    from st_aggrid import AgGrid, GridOptionsBuilder
    from st_aggrid.shared import GridUpdateMode
    from st_aggrid import DataReturnMode
    AGGRID_AVAILABLE = True
except Exception:
    AGGRID_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------
# Helpers and NSW utility functions
# ---------------------------

def detect_columns(df):
    lookup = {c.strip().upper(): c for c in df.columns}
    def pick(*names):
        for n in names:
            if n.upper() in lookup:
                return lookup[n.upper()]
        return None
    return {
        'PRODUCT': pick('Subject','PRODUCT','Item','ITEM','NAME'),
        'SCHEDULE BRAND': pick('Manufacturer','BRAND','MAKE','SCHEDULE BRAND'),
        'SCHEDULE MODEL': pick('Model','MODEL','MODEL #','SCHEDULE MODEL'),
        'BRAND': pick('Manufacturer','BRAND','MAKE'),
        'MODEL': pick('Model','MODEL','MODEL #'),
        'QTY': pick('Count','QTY','Quantity','QUANT'),
        'TAG': pick('Label','TAG','Mark'),
        'NECK SIZE': pick('Neck Size','NECK SIZE','NECKSIZE'),
        'MODULE SIZE': pick('Face Size','FACE','MODULE SIZE'),
        'DUCT SIZE': pick('Duct Size','DUCTSIZE'),
        'TYPE': pick('Type','DESCRIPTION','TYPE'),
        'MOUNTING': pick('Mounting','MOUNT'),
        'ACCESSORIES1': pick('Accessories','ACCESSORIES','ACCESSORIES1'),
        'ACCESSORIES2': pick('Description','DESCRIPTION'),
        'REMARK': pick('Remark','Remarks','Notes','REMARK')
    }

def pad_numeric_only(tag):
    s = str(tag).strip()
    return s.zfill(2) if re.fullmatch(r'\d+', s) else s

_token_re = re.compile(r'(\d+|\D+)')
def tag_sort_key(tag):
    t = str(tag).strip()
    if t == '':
        return (3, '', ())
    if re.match(r'^[A-Za-z]', t):
        parts = _token_re.findall(t)
        seq = tuple((0, p.lower()) if not p.isdigit() else (1, int(p)) for p in parts)
        return (0, t.lower(), seq)
    m = re.match(r'^(\d+)', t)
    if m:
        num = int(m.group(1))
        rest = t[m.end():].lower()
        return (1, num, rest)
    return (2, t.lower(), ())

def neck_size_key(val):
    m = re.search(r'(\d+(\.\d+)?)', str(val))
    return float(m.group(1)) if m else float('inf')

def mounting_rank(val):
    if not isinstance(val, str):
        return 2
    v = val.lower()
    lay_variants = ['lay-in','lay in','in lay','in laying','in-lay','layin','laying','inlay','lay']
    for variant in lay_variants:
        if variant in v:
            return 0
    if 'surface' in v:
        return 1
    return 2

def to_float_safe(x):
    try:
        s = str(x).strip()
        return float(s) if s!='' else 0.0
    except:
        return 0.0

_tag_prefix_re = re.compile(r'^([A-Za-z]+)')
def _tag_prefix(tag):
    m = _tag_prefix_re.match(str(tag).strip())
    return m.group(1).upper() if m else str(tag).strip().upper()

def summarize_tag_prefixes_with_hash(tags_series):
    prefixes = []
    for t in tags_series:
        t = str(t).strip()
        if t == '':
            continue
        prefixes.append(_tag_prefix(t))
    if not prefixes:
        return '', ' (TOTAL)'
    parts = [f"{p}#" for p in sorted(set(prefixes))]
    compact = "/".join(parts)
    return compact, f"{compact} (TOTAL)"

# Robust upload reader (returns (df, error_message))
def read_uploaded_file(uploaded_file):
    if uploaded_file is None:
        return None, "No file uploaded."

    name = getattr(uploaded_file, "name", str(uploaded_file)).lower()
    # Excel first
    if name.endswith(('.xls', '.xlsx')):
        try:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, dtype=str)
            if df.shape[1] == 0:
                return None, "Excel file read OK but no columns were detected."
            df = df.fillna('')
            df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
            return df, None
        except Exception as e:
            return None, f"Failed to read Excel file: {e}"

    # Otherwise treat as CSV-like
    try:
        uploaded_file.seek(0)
        raw = uploaded_file.read()
        if isinstance(raw, bytes):
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                try:
                    text = raw.decode('latin1')
                except Exception:
                    return None, "Failed to decode CSV file bytes as UTF-8 or Latin-1."
        else:
            text = str(raw)

        lines = [ln for ln in text.splitlines() if ln.strip() != '']
        if len(lines) == 0:
            return None, "CSV appears to be empty (no non-blank lines)."

        sample = "\n".join(lines[:10])

        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            delim = dialect.delimiter
        except Exception:
            for d in [',',';','\t','|']:
                if d in sample:
                    delim = d
                    break
            else:
                delim = ','

        try:
            df = pd.read_csv(io.StringIO(text), sep=delim, dtype=str, engine='python')
        except Exception as e:
            try:
                df = pd.read_csv(io.StringIO(text), sep=None, engine='python', dtype=str)
            except Exception as e2:
                return None, f"Failed to parse CSV: {e} (fallback error: {e2})"

        if df.shape[1] == 0:
            return None, "CSV read OK but no columns were detected."
        df = df.fillna('')
        df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
        return df, None

    except Exception as e:
        return None, f"Unexpected failure reading file: {e}"

# Styling & excel save (two sheets)
def style_and_save_two_sheets(source_df, takeoff_df, excel_bytes_io,
                         header_fill="#CFE2F3", subtotal_fill="#FFFF00", grand_fill="#CFE2F3",
                         bold_cols = ['PRODUCT','SCHEDULE MODEL','TAG','MODULE SIZE','TYPE','ACCESSORIES1']):
    wb = Workbook()
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # RawData sheet (first)
    ws_raw = wb.active
    ws_raw.title = "RawData"
    for c_idx, col in enumerate(source_df.columns, start=1):
        cell = ws_raw.cell(row=1, column=c_idx, value=col)
        cell.fill = PatternFill(start_color=header_fill.replace('#',''), end_color=header_fill.replace('#',''), fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    for r_idx, row in enumerate(source_df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_raw.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
    for i, col in enumerate(source_df.columns, 1):
        try:
            max_len = max(source_df[col].astype(str).map(len).max(), len(col)) + 2
        except Exception:
            max_len = len(col) + 2
        ws_raw.column_dimensions[get_column_letter(i)].width = max_len

    # TakeoffNSW sheet
    ws = wb.create_sheet(title="TakeoffNSW")
    for c_idx, col in enumerate(takeoff_df.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = PatternFill(start_color=header_fill.replace('#',''), end_color=header_fill.replace('#',''), fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for r_idx, row in enumerate(takeoff_df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            colname = takeoff_df.columns[c_idx-1]
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if colname in bold_cols and str(value).strip() != '':
                cell.font = Font(bold=True)
            if colname == 'QTY' and str(value).strip() != '':
                try:
                    tag_val = str(row[list(takeoff_df.columns).index('TAG')]).strip() if 'TAG' in takeoff_df.columns else ''
                except Exception:
                    tag_val = ''
                prod_val = str(row[0]).strip()
                if tag_val.endswith('(TOTAL)') or prod_val == 'Grand Total' or ('=' in prod_val and prod_val.split('=')[-1].strip().replace('.','',1).isdigit()):
                    cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=False)
            if colname == 'TAG' and isinstance(value, str) and value.endswith('(TOTAL)'):
                for name in ['PRODUCT','TAG','QTY']:
                    if name in takeoff_df.columns:
                        i = list(takeoff_df.columns).index(name) + 1
                        ccell = ws.cell(row=r_idx, column=i)
                        ccell.fill = PatternFill(start_color=subtotal_fill.replace('#',''), end_color=subtotal_fill.replace('#',''), fill_type="solid")
                        ccell.font = Font(bold=True)
                        ccell.border = border
            if value == 'Grand Total':
                for col_i in range(1, len(takeoff_df.columns)+1):
                    ccell = ws.cell(row=r_idx, column=col_i)
                    ccell.fill = PatternFill(start_color=grand_fill.replace('#',''), end_color=grand_fill.replace('#',''), fill_type="solid")
                    ccell.font = Font(bold=True)
                    ccell.border = border

    ws.freeze_panes = 'A2'
    for i, col in enumerate(takeoff_df.columns, 1):
        try:
            max_len = max(takeoff_df[col].astype(str).map(len).max(), len(col)) + 2
        except Exception:
            max_len = len(col) + 2
        ws.column_dimensions[get_column_letter(i)].width = max_len

    wb.save(excel_bytes_io)
    excel_bytes_io.seek(0)
    return excel_bytes_io

# ---------------------------
# Corrected build_takeoff (with manual tag normalization fix)
# ---------------------------

def build_takeoff(df_in,
                  mapping,
                  apply_brand_autofill=True,
                  keep_model_blank=True,
                  replace_empty_with_dot=False,
                  aggregate_duplicates=True,
                  product_order=None,
                  tag_order=None,
                  manual_tags_by_product=None):
    """
    Returns: src_df, final_df, integrity
    manual_tags_by_product: dict mapping normalized product -> list of tag strings (free-form). This function
    normalizes those saved tag strings (pad numeric-only to 2 digits and lowercase) to match df['TAG'] normalization.
    """
    target_cols = ['PRODUCT','SCHEDULE BRAND','SCHEDULE MODEL','BRAND','MODEL','QTY','TAG',
                   'NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2','REMARK']

    # Map columns -> this will be our source (RawData)
    src_df = pd.DataFrame()
    for c in target_cols:
        src = mapping.get(c)
        if src and src in df_in.columns:
            src_df[c] = df_in[src].astype(str).fillna('')
        else:
            src_df[c] = ''

    # Working copy
    df = src_df.copy()

    # Brand autofill
    if apply_brand_autofill:
        prod_upper = df['PRODUCT'].astype(str).str.upper()
        df['BRAND'] = df.get('BRAND','').astype(str)
        mask_adgrd = prod_upper.str.contains('AD-GRD', na=False) & (df['BRAND'].astype(str).str.strip()=='')
        mask_fan = prod_upper.str.contains('FAN', na=False) & (df['BRAND'].astype(str).str.strip()=='')
        mask_split = prod_upper.str.contains('SPLIT SYSTEM', na=False) & (df['BRAND'].astype(str).str.strip()=='')
        df.loc[mask_adgrd, 'BRAND'] = 'PRICE'
        df.loc[mask_fan, 'BRAND'] = 'LOREN COOK'
        df.loc[mask_split, 'BRAND'] = 'SAMSUNG'
    else:
        if mapping.get('BRAND') and mapping['BRAND'] in df_in.columns:
            df['BRAND'] = df_in[mapping['BRAND']].astype(str).fillna('')
        else:
            df['BRAND'] = ''

    # MODEL blank handling
    if keep_model_blank:
        df['MODEL'] = ''
    else:
        if mapping.get('MODEL') and mapping['MODEL'] in df_in.columns:
            df['MODEL'] = df_in[mapping['MODEL']].astype(str).fillna('')

    # Replace empties with dot if requested
    if replace_empty_with_dot:
        df = df.replace('', '.').replace('nan','.')

    # numeric qty and tag normalization
    df['QTY'] = df['QTY'].apply(to_float_safe)
    df['TAG'] = df['TAG'].astype(str).apply(lambda t: pad_numeric_only(t.strip()))

    # optional aggregation of exact duplicate attribute rows
    if aggregate_duplicates:
        agg_keys = ['PRODUCT','TAG','SCHEDULE BRAND','SCHEDULE MODEL','NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2']
        for k in agg_keys:
            if k not in df.columns:
                df[k] = ''
        grouped = df.groupby(agg_keys, dropna=False, as_index=False).agg({
            'BRAND': 'first',
            'MODEL': 'first',
            'QTY': 'sum',
            'REMARK': lambda x: '; '.join([v for v in x.astype(str).unique() if v and v!='nan'])
        })
        new_df = pd.DataFrame()
        for c in target_cols:
            new_df[c] = grouped.get(c, '')
        df = new_df

    # ordering keys
    def _display_name_for_order(x, blank_placeholder):
        s = '' if (pd.isna(x) or str(x).strip() == '') else str(x).strip()
        return blank_placeholder if s == '' else s

    if product_order:
        prod_norm_order = [_display_name_for_order(p, '<<BLANK PRODUCT>>') for p in product_order]
        prod_rank = {p: i for i, p in enumerate(prod_norm_order)}
        df['_prod_key'] = df['PRODUCT'].apply(lambda x: prod_rank.get(_display_name_for_order(x, '<<BLANK PRODUCT>>'), len(prod_rank)))
    else:
        df['_prod_key'] = df['PRODUCT'].fillna('').astype(str).str.lower()

    # normalize manual_tags_by_product into manual_by_product using same normalization as df['TAG']
    manual_by_product = {}
    if manual_tags_by_product and isinstance(manual_tags_by_product, dict):
        for k, v in manual_tags_by_product.items():
            if not isinstance(v, list):
                continue
            pnorm = str(k).strip().lower()
            normalized_tags = []
            for x in v:
                if x is None:
                    continue
                xstr = str(x).strip()
                xnorm = pad_numeric_only(xstr)
                xnorm = xnorm.lower()
                normalized_tags.append(xnorm)
            manual_by_product[pnorm] = normalized_tags

    global_tag_rank = {}
    if tag_order and isinstance(tag_order, list):
        normalized = [str(t).strip() for t in tag_order]
        global_tag_rank = {t: i for i, t in enumerate(normalized)}

    def compute_tag_key(row):
        prod = str(row.get('PRODUCT','')).strip().lower()
        raw_tag = str(row.get('TAG','')).strip()
        tag_norm = pad_numeric_only(raw_tag).lower()
        L = manual_by_product.get(prod)
        if L:
            if tag_norm in L:
                return ('manual', L.index(tag_norm))
            else:
                return ('manual_auto', tag_sort_key(raw_tag))
        if raw_tag in global_tag_rank:
            return ('global', global_tag_rank[raw_tag])
        return ('auto', tag_sort_key(raw_tag))

    df['_tag_key'] = df.apply(lambda r: compute_tag_key(r), axis=1)
    df['_neck'] = df['NECK SIZE'].apply(neck_size_key)
    df['_mount'] = df['MOUNTING'].apply(mounting_rank)

    # final sort
    df = df.sort_values(by=['_prod_key','_tag_key','_neck','_mount'], kind='mergesort').reset_index(drop=True)

    # build output rows
    out_rows = []
    grand_total = 0.0
    input_tags = set(df['TAG'].unique())

    for prod_key, group in df.groupby('_prod_key', sort=True):
        group = group.copy().reset_index(drop=True)
        if group.empty:
            continue
        prod_name = group.loc[0,'PRODUCT']
        is_adgrd = isinstance(prod_name, str) and 'AD-GRD' in prod_name.upper()
        if is_adgrd:
            first_of_product = True
            tags = group['TAG'].unique().tolist()
            for tag in tags:
                tag_group = group[group['TAG'] == tag]
                subtotal = tag_group['QTY'].sum()
                grand_total += subtotal
                for idx, r in tag_group.iterrows():
                    row = {c: r[c] if c in r.index else '' for c in target_cols}
                    row['PRODUCT'] = prod_name if first_of_product else ''
                    out_rows.append(row)
                    first_of_product = False
                compact, tagcol = summarize_tag_prefixes_with_hash(tag_group['TAG'])
                num_display = int(subtotal) if subtotal == int(subtotal) else float(f"{subtotal:.2f}")
                subtotal_row = {c: '' for c in target_cols}
                subtotal_row['PRODUCT'] = (f"{num_display} = {compact}" if compact else f"{tag}={num_display}")
                subtotal_row['TAG'] = (f"{compact} (TOTAL)" if compact else f"{tag} (TOTAL)")
                subtotal_row['QTY'] = subtotal
                out_rows.append(subtotal_row)
            out_rows.append({**{c:'' for c in target_cols}})
        else:
            first_of_product = True
            prod_total = 0.0
            for idx, r in group.iterrows():
                row = {c: r[c] if c in r.index else '' for c in target_cols}
                row['PRODUCT'] = prod_name if first_of_product else ''
                out_rows.append(row)
                first_of_product = False
                prod_total += r['QTY']
                grand_total += r['QTY']
            compact, tagcol = summarize_tag_prefixes_with_hash(group['TAG'])
            num_display = int(prod_total) if prod_total == int(prod_total) else float(f"{prod_total:.2f}")
            subtotal_row = {c: '' for c in target_cols}
            if compact:
                subtotal_row['PRODUCT'] = f"{num_display} = {compact}"
                subtotal_row['TAG'] = f"{compact} (TOTAL)"
            else:
                subtotal_row['PRODUCT'] = f"{prod_name}={int(prod_total) if prod_total==int(prod_total) else '{:.2f}'.format(prod_total)}"
                subtotal_row['TAG'] = f"{prod_name.split('=')[0] if '=' in prod_name else prod_name.split('=')[0]} (TOTAL)"
            subtotal_row['QTY'] = prod_total
            out_rows.append(subtotal_row)
            out_rows.append({**{c:'' for c in target_cols}})

    grand_row = {c: '' for c in target_cols}
    grand_row['PRODUCT'] = 'Grand Total'
    grand_row['TAG'] = 'Grand Total'
    grand_row['QTY'] = grand_total
    out_rows.append(grand_row)

    final_df = pd.DataFrame(out_rows, columns=target_cols)

    # integrity
    output_tags = set([t for t in final_df['TAG'].dropna() if isinstance(t,str) and t.strip()!='' and not t.endswith('(TOTAL)')])
    missing_tags = sorted(list(input_tags - output_tags))
    integrity = {
        "input_unique_tags": len(input_tags),
        "output_unique_tags": len([t for t in final_df['TAG'].unique() if t and not str(t).endswith('(TOTAL)')]),
        "all_tags_preserved": len(missing_tags) == 0,
        "missing_tags_sample": missing_tags[:20]
    }

    return src_df, final_df, integrity

# ---------------------------
# Small helpers used by Triune UI (normalize product text, ordering)
# ---------------------------

def normalize_product_text(s: str) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    return t.lower()

def _norm_key(s: str) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()

def safe_rerun():
    try:
        rerun_fn = getattr(st, "experimental_rerun", None)
        if callable(rerun_fn):
            rerun_fn()
            return
        try:
            from streamlit.runtime.scriptrunner import RerunException
            raise RerunException()
        except Exception:
            return
    except Exception:
        return

# ---------------------------
# Manual-only Editors (textarea + Apply)
# ---------------------------

def render_manual_product_editor(preview_df: pd.DataFrame):
    product_values = preview_df["PRODUCT"].fillna("").astype(str).tolist()
    seen = set()
    products = []
    for p in product_values:
        if p not in seen:
            seen.add(p)
            products.append(p)

    if not products:
        st.info("No products found to order.")
        return

    with st.expander("Manual PRODUCT ordering — one product per line (click Apply to save)", expanded=False):
        fallback_key = "manual_order_textarea"
        default_text = "\n".join(products)
        pref = st.session_state.get('triune_manual_products_original')
        if pref and isinstance(pref, list) and len(pref) > 0:
            default_text = "\n".join(pref)
        pasted = st.text_area("Manual product order (one product per line):", value=st.session_state.get(fallback_key, default_text), height=240, key=fallback_key)
        lines = [ln.strip() for ln in pasted.splitlines() if ln.strip() != ""]
        preview_order = lines or products
        st.markdown("**Preview (unsaved) product order:**")
        st.write(preview_order)

        manual_norm = []
        manual_orig = []
        seen2 = set()
        for orig in preview_order:
            pnorm = normalize_product_text(orig)
            if pnorm != "" and pnorm not in seen2:
                seen2.add(pnorm)
                manual_norm.append(pnorm)
                manual_orig.append(orig)

        col1, col2 = st.columns([1,1])
        if col1.button("Apply product order"):
            if manual_norm:
                st.session_state['triune_manual_products'] = manual_norm
                st.session_state['triune_manual_products_original'] = manual_orig
                st.success(f"Product order applied — {len(manual_norm)} products saved.")
                safe_rerun()
            else:
                st.session_state.pop('triune_manual_products', None)
                st.session_state.pop('triune_manual_products_original', None)
                st.info("Cleared manual product order (nothing saved).")
                safe_rerun()

        if col2.button("Reset manual product order"):
            st.session_state.pop('triune_manual_products', None)
            st.session_state.pop('triune_manual_products_original', None)
            st.success("Manual product order cleared.")
            safe_rerun()

def render_manual_per_product_tag_editor(preview_df: pd.DataFrame):
    detected_products = preview_df["PRODUCT"].fillna("").astype(str).tolist()
    seen = set(); prod_list = []
    for p in detected_products:
        pn = _norm_key(p)
        if pn not in seen:
            seen.add(pn); prod_list.append(p)

    manual_prod_norm = st.session_state.get('triune_manual_products') or []
    manual_prod_display = []
    if manual_prod_norm:
        orig_map = {}
        for orig in prod_list:
            orig_map[_norm_key(orig)] = orig
        for mp in manual_prod_norm:
            if mp in orig_map:
                manual_prod_display.append(orig_map[mp])
        for p in prod_list:
            if p not in manual_prod_display:
                manual_prod_display.append(p)
        products_for_ui = manual_prod_display
    else:
        products_for_ui = prod_list

    if not products_for_ui:
        st.info("No products found for per-product tag editor.")
        return

    prev = st.session_state.get('triune_perprod_selected_product')
    sel_index = 0
    if prev and prev in products_for_ui:
        sel_index = products_for_ui.index(prev)
    sel_prod = st.selectbox("Select product to edit tags for", options=products_for_ui, index=sel_index, key="perprod_tag_select")
    st.session_state['triune_perprod_selected_product'] = sel_prod
    sel_prod_norm = _norm_key(sel_prod)

    # collect tags for selected product (unique order-preserving)
    tags_raw = preview_df.loc[preview_df["PRODUCT"].fillna("").astype(str) == sel_prod, "TAG"].fillna("").astype(str).tolist()
    seen_t = set(); tags = []
    for t in tags_raw:
        if t not in seen_t:
            seen_t.add(t); tags.append(t)

    if not tags:
        st.info(f"No TAGs found for product: {sel_prod}")
        return

    with st.expander(f"Edit TAG order for product: {sel_prod} (click Apply to save)", expanded=False):
        fallback_key = f"manual_tag_order_{_norm_key(sel_prod)}"
        default_text = "\n".join(tags)
        saved_orig_map = st.session_state.get('triune_manual_tags_by_product_orig') or {}
        if sel_prod_norm in saved_orig_map:
            default_text = "\n".join(saved_orig_map[sel_prod_norm])
        pasted = st.text_area("Manual tag order for this product (one tag per line):", value=st.session_state.get(fallback_key, default_text), height=200, key=fallback_key)
        lines = [ln.strip() for ln in pasted.splitlines() if ln.strip() != ""]
        preview_order = lines or tags
        st.markdown("**Preview (unsaved) tag order for product:**")
        st.write(preview_order)

        manual_norm_list = []
        manual_orig_list = []
        for tv in preview_order:
            tn = _norm_key(tv)
            if tn != "":
                manual_norm_list.append(tn)
                manual_orig_list.append(tv)

        col1, col2 = st.columns([1,1])
        if col1.button(f"Apply TAG order for product: {sel_prod}"):
            if 'triune_manual_tags_by_product' not in st.session_state:
                st.session_state['triune_manual_tags_by_product'] = {}
            if 'triune_manual_tags_by_product_orig' not in st.session_state:
                st.session_state['triune_manual_tags_by_product_orig'] = {}
            if manual_norm_list:
                st.session_state['triune_manual_tags_by_product'][sel_prod_norm] = manual_norm_list
                st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm] = manual_orig_list
                st.success(f"Saved manual TAG order for product: {sel_prod} ({len(manual_norm_list)} tags).")
                safe_rerun()
            else:
                if sel_prod_norm in st.session_state.get('triune_manual_tags_by_product', {}):
                    del st.session_state['triune_manual_tags_by_product'][sel_prod_norm]
                if sel_prod_norm in st.session_state.get('triune_manual_tags_by_product_orig', {}):
                    del st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm]
                st.info(f"Cleared manual TAG order for product: {sel_prod} (nothing saved).")
                safe_rerun()

        if col2.button(f"Reset TAG order for product: {sel_prod}"):
            if 'triune_manual_tags_by_product' in st.session_state and sel_prod_norm in st.session_state['triune_manual_tags_by_product']:
                del st.session_state['triune_manual_tags_by_product'][sel_prod_norm]
            if 'triune_manual_tags_by_product_orig' in st.session_state and sel_prod_norm in st.session_state['triune_manual_tags_by_product_orig']:
                del st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm]
            st.success(f"Cleared saved TAG order for product: {sel_prod}.")
            safe_rerun()

# ---------------------------
# Triune Data Unit helper functions
# ---------------------------

def normalize_text(s: str) -> str:
    return re.sub(r"[\s_\-\.]+", " ", str(s).strip().lower())

def find_column(raw_cols: List[str], *aliases: str) -> Optional[str]:
    raw_norm = {c: normalize_text(c) for c in raw_cols}
    for alias in aliases:
        a_norm = normalize_text(alias)
        for rc, rn in raw_norm.items():
            if rn == a_norm:
                return rc
    for alias in aliases:
        a_norm = normalize_text(alias)
        for rc, rn in raw_norm.items():
            if a_norm in rn or rn in a_norm:
                return rc
    alias_tokens = set()
    for alias in aliases:
        alias_tokens.update(normalize_text(alias).split())
    best = None
    best_score = 0
    for rc, rn in raw_norm.items():
        tokens = set(rn.split())
        score = len(tokens & alias_tokens)
        if score > best_score:
            best_score = score
            best = rc
    if best_score > 0:
        return best
    return None

def load_file_to_df_simple_for_units(uploaded_file):
    if uploaded_file is None:
        return None
    df, err = read_uploaded_file(uploaded_file)
    if err:
        return None
    return df

def detect_unit_column(df: pd.DataFrame) -> Optional[str]:
    for c in df.columns:
        if re.search(r"(^unit$)|unit(s)?|apt|apartment|flat|room|suite|unit_id|unitno|unit_no", c, re.IGNORECASE):
            return c
    return None

def apply_raw_column_mapping(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = raw_df.copy()
    cols = list(df.columns)
    lowered = {c.lower(): c for c in cols}
    exact_rules = {
        "subject": "PRODUCT",
        "page index": "Page Index",
        "label": "TAG",
        "manufacturer": "BRAND",
        "face size": "MODULE SIZE",
        "description": "ACCESSORIES2",
        "accessories": "ACCESSORIES1",
        "accessories1": "ACCESSORIES1"
    }
    exact_map = {}
    for low_key, dest in exact_rules.items():
        if low_key in lowered:
            orig = lowered[low_key]
            if orig != dest:
                exact_map[orig] = dest
    if exact_map:
        df = df.rename(columns=exact_map)
        cols = list(df.columns)
    mapping_pairs = {
        "QTY": ("quantity", "qty", "count", "qty.", "q'ty"),
        "BRAND": ("brand", "make", "mfr"),
        "MODEL": ("model", "catalog", "cat no", "catalog no"),
        "TAG": ("tag", "tag id", "label", "ref", "mark"),
        "NECK SIZE": ("neck size", "neck"),
        "MODULE SIZE": ("module size", "face size", "module", "face"),
        "DUCT SIZE": ("duct size", "duct"),
        "CFM": ("cfm", "airflow"),
        "TYPE": ("type", "desc", "description"),
        "MOUNTING": ("mounting", "install", "mount"),
        "ACCESSORIES1": ("accessories", "accessories1", "accessory 1", "accessory"),
        "ACCESSORIES2": ("accessories2", "accessory 2", "description", "desc"),
        "REMARK": ("remark", "remarks", "note", "notes"),
        "UNITS": ("unit", "units", "zone", "area", "apt", "apartment")
    }
    rename_dict = {}
    cols = list(df.columns)
    for target_col, aliases in mapping_pairs.items():
        if target_col in df.columns:
            continue
        found = find_column(cols, *aliases)
        if found and found not in rename_dict and found != target_col:
            rename_dict[found] = target_col
    if rename_dict:
        df = df.rename(columns=rename_dict)
    return df

def clean_unit_matrix(df: pd.DataFrame, unit_col_hint: Optional[str] = None) -> (pd.DataFrame, str):
    df2 = df.copy()
    unit_col = unit_col_hint or detect_unit_column(df2) or df2.columns[0]
    df2[unit_col] = df2[unit_col].fillna("").astype(str).str.strip()
    df2 = df2[~df2[unit_col].str.upper().isin(["TOTAL", "GRAND TOTAL", "SUMMARY", "ALL"])]
    return df2.reset_index(drop=True), unit_col

def guess_multiplier_column(unit_df: pd.DataFrame, unit_col: str) -> Optional[str]:
    candidates = [c for c in unit_df.columns if c != unit_col]
    for c in candidates:
        sample = unit_df[c].dropna().astype(str).str.replace(",", "").str.strip()
        if sample.size and any(s.replace(".", "", 1).isdigit() for s in sample[:50]):
            return c
    return candidates[0] if candidates else None

def build_data_unit_sheet(
    raw_df: pd.DataFrame,
    unit_df: pd.DataFrame,
    raw_unit_col: str,
    matrix_unit_col: str,
    multiplier_col: Optional[str],
    selected_units: Optional[list] = None,
    include_empty: bool = True,
    default_multiplier: int = 1
) -> pd.DataFrame:
    raw = raw_df.copy()
    mat = unit_df.copy()
    raw[raw_unit_col] = raw[raw_unit_col].fillna("").astype(str).str.strip()
    mat[matrix_unit_col] = mat[matrix_unit_col].fillna("").astype(str).str.strip()
    multiplier_map = {}
    if multiplier_col:
        for _, r in mat.iterrows():
            u = str(r.get(matrix_unit_col, "")).strip()
            v = r.get(multiplier_col, "")
            try:
                if pd.isna(v) or str(v).strip() == "":
                    continue
                num = int(float(str(v).replace(",", "").strip()))
                multiplier_map[u] = num
            except Exception:
                continue
    org_count_col = None
    for name in ["Org Count", "OrgCount", "ORIGINAL COUNT", "Count", "QTY", "Qty", "qty", "QTY/UNIT"]:
        if name in raw.columns:
            org_count_col = name
            break
    if org_count_col is None:
        for c in raw.columns:
            sample = raw[c].dropna().astype(str).str.replace(",", "").str.strip()
            if sample.size and all(s.replace(".", "", 1).isdigit() for s in sample[:50]):
                org_count_col = c
                break
    if org_count_col is None:
        raw["Org Count"] = 1
    else:
        raw["Org Count"] = pd.to_numeric(raw[org_count_col].fillna("0").astype(str).str.replace(",", "").str.strip(), errors="coerce").fillna(0).astype(int)
    def get_mult(u):
        if u == "" or pd.isna(u):
            return default_multiplier
        return multiplier_map.get(u, default_multiplier)
    raw["__unit_multiplier__"] = raw[raw_unit_col].apply(get_mult).astype(int)
    raw["Count"] = raw["Org Count"].astype(int) * raw["__unit_multiplier__"]
    if selected_units:
        sel = set(selected_units)
        raw_units_vals = raw[raw_unit_col].fillna("").astype(str).str.strip()
        mask = raw_units_vals.isin(sel)
        if "<<EMPTY UNIT>>" in sel:
            mask = mask | (raw_units_vals == "")
        raw = raw[mask].copy()
    out_unit_name = "UNITS"
    raw[out_unit_name] = raw[raw_unit_col].replace({"": "<<EMPTY UNIT>>"})
    required_order = [
        "PRODUCT", "Page Index", "TAG", "Org Count", "Count", "BRAND", "MODEL",
        "NECK SIZE", "MODULE SIZE", "DUCT SIZE", "CFM", "TYPE", "MOUNTING",
        "ACCESSORIES1", "ACCESSORIES2", "REMARK", out_unit_name, "DAMPER TYPE"
    ]
    out = raw.copy()
    for col in required_order:
        if col not in out.columns:
            out[col] = ""
    other_cols = [c for c in out.columns if c not in required_order and c != "__unit_multiplier__"]
    final_cols = required_order + other_cols
    if "__unit_multiplier__" in out.columns:
        out = out.drop(columns=["__unit_multiplier__"])
    return out[final_cols].copy()

def bytes_from_df_excel(df: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data Unit Sheet")
    return out.getvalue()

def create_zip_bytes_from_map(dfs_map: dict, export_fn) -> bytes:
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for unit, df in dfs_map.items():
            safe = "".join(c if c.isalnum() or c in " -_." else "_" for c in str(unit))[:120] or "unit"
            fname = f"{safe}.xlsx"
            zf.writestr(fname, export_fn(df))
    return mem.getvalue()

def export_styled_excel_bytes(df: pd.DataFrame) -> bytes:
    return bytes_from_df_excel(df)

# ---------------------------
# Streamlit UI: Two tabs — Takeoff & Data Units
# ---------------------------

APP_TITLE = "TakeoffNSW + Data Unit Tools (Manual ordering)"

st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)
st.markdown("Upload raw takeoff and convert to TakeoffNSW format (Triune-compatible), or build Data Unit Sheets (Triune helpers). Manual PRODUCT ordering and per-product TAG ordering included.")

tabs = st.tabs(["Takeoff", "Data Units"])

# ---------- TAB: Takeoff ----------
with tabs[0]:
    st.header("TakeoffNSW Formatter (Triune-compatible output)")
    uploaded = st.file_uploader("Upload CSV or Excel (raw takeoff)", type=['csv','xlsx','xls'], key="nsw_upload")

    st.sidebar.header("Options")
    apply_brand = st.sidebar.checkbox("Apply brand autofill rules (AD-GRD/FAN/SPLIT SYSTEM)", value=True)
    keep_model_blank = st.sidebar.checkbox("Keep MODEL blank", value=True)
    replace_empty = st.sidebar.checkbox("Replace empty cells with '.'", value=False)
    aggregate_duplicates = st.sidebar.checkbox("Aggregate exact duplicate attribute rows (sum QTY)", value=True)

    st.sidebar.markdown("### Colors")
    header_color = st.sidebar.color_picker("Header fill", "#CFE2F3")
    subtotal_color = st.sidebar.color_picker("Subtotal fill", "#FFFF00")
    grand_color = st.sidebar.color_picker("Grand total fill", "#CFE2F3")

    st.sidebar.markdown("### Bold columns (comma-separated)")
    bold_input = st.sidebar.text_input("Columns to bold (exact names)", value="PRODUCT,SCHEDULE MODEL,TAG,MODULE SIZE,TYPE,ACCESSORIES1")
    bold_cols = [c.strip() for c in bold_input.split(',') if c.strip()]

    st.sidebar.markdown("### Mapping JSON (optional)")
    mapping_json = st.sidebar.text_area("Mapping JSON (target -> source). Leave empty to auto-detect.", value="", height=140, key="nsw_mapjson")
    if st.sidebar.button("Load mapping from JSON", key="nsw_loadmapbtn"):
        try:
            m = json.loads(mapping_json)
            st.session_state['nsw_mapping'] = m
            st.sidebar.success("Mapping loaded into session.")
        except Exception as e:
            st.sidebar.error(f"Invalid JSON: {e}")

    if uploaded:
        df_in, read_err = read_uploaded_file(uploaded)
        if read_err:
            st.error(f"Failed to read file: {read_err}")
        else:
            auto_map = detect_columns(df_in)
            mapping_use = auto_map.copy()
            if st.session_state.get('nsw_mapping'):
                for k,v in st.session_state['nsw_mapping'].items():
                    mapping_use[k] = v

            st.subheader("Column mapping (target → source)")
            with st.form("mapping_form_nsw"):
                cols = ['PRODUCT','SCHEDULE BRAND','SCHEDULE MODEL','BRAND','MODEL','QTY','TAG','NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2','REMARK']
                user_map = {}
                for c in cols:
                    choices = [''] + list(df_in.columns)
                    default = mapping_use.get(c) if mapping_use.get(c) in df_in.columns else ''
                    user_map[c] = st.selectbox(c, choices, index=(choices.index(default) if default in choices else 0), key=f"map_{c}_nsw")
                save_map = st.form_submit_button("Apply mapping")
            if save_map:
                st.session_state['nsw_mapping'] = user_map
                st.success("Mapping applied and saved.")

            mapping_use_final = st.session_state.get('nsw_mapping') or mapping_use

            # Preview mapped source
            preview_build = pd.DataFrame()
            for c in cols:
                src = (mapping_use_final.get(c) if mapping_use_final.get(c) else auto_map.get(c))
                if src and src in df_in.columns:
                    preview_build[c] = df_in[src].astype(str).fillna('')
                else:
                    preview_build[c] = ''

            st.markdown("**Preview of mapped data (first 20 rows)**")
            st.dataframe(preview_build.head(20), height=300)

            # Manual product & per-product tag editors
            st.markdown("### Manual PRODUCT ordering")
            st.caption("One product per line. Click Apply to save a manual product order which will be used when generating the takeoff.")
            render_manual_product_editor(preview_build)

            st.markdown("### Per-product TAG ordering")
            st.caption("Select a product, then paste one tag per line in the textarea and click Apply. This per-product ordering overrides global tag sort for that product.")
            render_manual_per_product_tag_editor(preview_build)

            st.markdown("---")
            st.markdown("This app uses automatic product/tag ordering **unless** you supply a manual product list (Apply) or per-product tag lists (Apply).")

            if st.button("Generate Takeoff files (preview + download)", key="gen_takeoff"):
                with st.spinner("Building takeoff..."):
                    prod_ord = st.session_state.get('triune_manual_products_original') if st.session_state.get('triune_manual_products_original') else None
                    manual_tags = st.session_state.get('triune_manual_tags_by_product') or {}

                    src_df, final_df, integrity = build_takeoff(
                        df_in,
                        mapping_use_final,
                        apply_brand_autofill=apply_brand,
                        keep_model_blank=keep_model_blank,
                        replace_empty_with_dot=replace_empty,
                        aggregate_duplicates=aggregate_duplicates,
                        product_order=prod_ord,
                        tag_order=None,
                        manual_tags_by_product=manual_tags
                    )

                    st.subheader("Integrity check")
                    st.json(integrity)

                    st.subheader("Final Takeoff preview (first 200 rows)")
                    st.dataframe(final_df.head(200), height=400)

                    excel_io = io.BytesIO()
                    style_and_save_two_sheets(src_df, final_df, excel_io,
                                         header_fill=header_color, subtotal_fill=subtotal_color, grand_fill=grand_color,
                                         bold_cols=bold_cols)

                    csv_io = io.BytesIO()
                    csv_io.write(final_df.to_csv(index=False).encode('utf-8'))
                    csv_io.seek(0)

                    st.download_button("📥 Download Triune Takeoff Excel Output (RawData + Takeoff)", data=excel_io,
                                       file_name=f"Triune_Takeoff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    st.download_button("📥 Download CSV (Takeoff only)", data=csv_io,
                                       file_name=f"Triune_Takeoff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

                    st.success("Files generated. Download above.")
    else:
        st.info("Upload a CSV or Excel file to begin the Takeoff conversion.")

# ---------- TAB: Data Units ----------
with tabs[1]:
    st.header("Data Unit Sheet Builder (Triune helpers)")
    col1, col2 = st.columns(2)
    with col1:
        raw_file = st.file_uploader("Upload Raw Takeoff (CSV or Excel)", type=["csv", "xlsx", "xls"], key="du_raw")
    with col2:
        unit_file = st.file_uploader("Upload Unit Matrix (CSV or Excel)", type=["csv", "xlsx", "xls"], key="du_matrix")

    if not raw_file or not unit_file:
        st.info("Upload both Raw Takeoff and Unit Matrix files to continue.")
    else:
        raw_df = load_file_to_df_simple_for_units(raw_file)
        unit_df = load_file_to_df_simple_for_units(unit_file)
        if raw_df is None or unit_df is None:
            st.error("Couldn't read one of the uploaded files. Please check file formats/encoding.")
        else:
            raw_df = apply_raw_column_mapping(raw_df)
            raw_unit_detected = detect_unit_column(raw_df)
            unit_matrix_clean, matrix_unit_detected = clean_unit_matrix(unit_df)

            st.markdown("### Detected columns")
            st.write(f"Raw unit column (auto-detected): **{raw_unit_detected or 'None'}**")
            st.write(f"Unit matrix unit column (auto-detected): **{matrix_unit_detected}**")

            raw_unit_col = st.selectbox("Select unit column in raw file", options=list(raw_df.columns),
                                        index=list(raw_df.columns).index(raw_unit_detected) if raw_unit_detected in raw_df.columns else 0,
                                        key="du_raw_unit_col")
            matrix_unit_col = st.selectbox("Select unit column in unit matrix", options=list(unit_matrix_clean.columns),
                                           index=list(unit_matrix_clean.columns).index(matrix_unit_detected) if matrix_unit_detected in unit_matrix_clean.columns else 0,
                                           key="du_matrix_unit_col")

            guessed_mult = guess_multiplier_column(unit_matrix_clean, matrix_unit_col)
            mult_options = [None] + list(unit_matrix_clean.columns)
            mult_index = 1 + list(unit_matrix_clean.columns).index(guessed_mult) if guessed_mult in unit_matrix_clean.columns else 0
            multiplier_col = st.selectbox("Select multiplier column in unit matrix (units per type)", options=mult_options,
                                         index=mult_index, key="du_multcol")
            if multiplier_col is None:
                st.warning("No multiplier column selected — multipliers default to 1 when unmatched.")

            units_series = unit_matrix_clean[matrix_unit_col].dropna().astype(str).str.strip()
            raw_units = raw_df[raw_unit_col].fillna("").astype(str).str.strip().unique().tolist()
            has_empty = any(u == "" for u in raw_df[raw_unit_col].fillna("").astype(str).str.strip().tolist())
            units_list = sorted(set(units_series.tolist() + [u for u in raw_units if u != ""]))
            if has_empty:
                units_list = ["<<EMPTY UNIT>>"] + units_list

            st.markdown("### Choose units to include (checkboxes)")
            select_all = st.checkbox("Select all units", value=True, key="du_selall")
            cols_chk = st.columns(3)
            selected_units = []
            for i, unit in enumerate(units_list):
                col = cols_chk[i % 3]
                default_val = True if select_all else False
                checked = col.checkbox(str(unit), value=default_val, key=f"du_unit_chk_{i}")
                if checked:
                    selected_units.append(unit)

            include_empty = st.checkbox("Include empty-unit rows (as <<EMPTY UNIT>>)", value=True, key="du_inc_empty")
            split_by_unit = st.checkbox("Split into separate Excel per unit (download as ZIP)", value=False, key="du_splitzip")

            st.markdown("### Output filename")
            default_name = "Data_Unit_Sheet"
            if split_by_unit:
                default_name = "takeoff_by_unit"
            file_name_input = st.text_input("Enter desired download filename (without extension):", value=default_name, key="du_outfilebase")
            def sanitize_filename_local(s: str) -> str:
                return "".join(c for c in s if c.isalnum() or c in " -_").strip() or default_name
            out_file_base = sanitize_filename_local(file_name_input)

            if st.button("Generate Data Unit Sheet", key="gen_du"):
                with st.spinner("Building Data Unit Sheet..."):
                    sel_units = selected_units if selected_units else None
                    final_df = build_data_unit_sheet(
                        raw_df=raw_df,
                        unit_df=unit_matrix_clean,
                        raw_unit_col=raw_unit_col,
                        matrix_unit_col=matrix_unit_col,
                        multiplier_col=multiplier_col,
                        selected_units=sel_units,
                        include_empty=include_empty,
                        default_multiplier=1
                    )
                    st.success("Data Unit Sheet created.")
                    st.write("Preview (first 200 rows):")
                    st.dataframe(final_df.head(200), use_container_width=True)

                    if split_by_unit:
                        groups = {u: g for u, g in final_df.groupby("UNITS")}
                        st.write("Files to be included in ZIP:")
                        summary = [{"unit": k, "rows": len(v)} for k, v in groups.items()]
                        st.table(pd.DataFrame(summary).sort_values("rows", ascending=False))
                        zip_bytes = create_zip_bytes_from_map(groups, export_styled_excel_bytes)
                        download_name = f"{out_file_base}.zip"
                        st.download_button("Download ZIP (per-unit Excels)", data=zip_bytes, file_name=download_name, mime="application/zip")
                    else:
                        excel_bytes = export_styled_excel_bytes(final_df)
                        download_name = f"{out_file_base}.xlsx"
                        st.download_button("Download combined Excel", data=excel_bytes, file_name=download_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
# takeoffnsw_with_units_manual_ordering.py
"""
Single-file Streamlit app:
- TakeoffNSW Formatter (Triune-compatible output)
- Data Unit Sheet builder (Triune helpers)
- Manual PRODUCT ordering (textarea + Apply) and Per-product TAG ordering (textarea + Apply)
Save and run with:
    streamlit run takeoffnsw_with_units_manual_ordering.py
"""

import io
import re
import json
import csv
import zipfile
from datetime import datetime
from typing import Optional, Tuple, Dict, List

import pandas as pd
import numpy as np
import streamlit as st

# optional st_aggrid (kept safe, not required)
try:
    from st_aggrid import AgGrid, GridOptionsBuilder
    from st_aggrid.shared import GridUpdateMode
    from st_aggrid import DataReturnMode
    AGGRID_AVAILABLE = True
except Exception:
    AGGRID_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------
# Helpers and NSW utility functions
# ---------------------------

def detect_columns(df):
    lookup = {c.strip().upper(): c for c in df.columns}
    def pick(*names):
        for n in names:
            if n.upper() in lookup:
                return lookup[n.upper()]
        return None
    return {
        'PRODUCT': pick('Subject','PRODUCT','Item','ITEM','NAME'),
        'SCHEDULE BRAND': pick('Manufacturer','BRAND','MAKE','SCHEDULE BRAND'),
        'SCHEDULE MODEL': pick('Model','MODEL','MODEL #','SCHEDULE MODEL'),
        'BRAND': pick('Manufacturer','BRAND','MAKE'),
        'MODEL': pick('Model','MODEL','MODEL #'),
        'QTY': pick('Count','QTY','Quantity','QUANT'),
        'TAG': pick('Label','TAG','Mark'),
        'NECK SIZE': pick('Neck Size','NECK SIZE','NECKSIZE'),
        'MODULE SIZE': pick('Face Size','FACE','MODULE SIZE'),
        'DUCT SIZE': pick('Duct Size','DUCTSIZE'),
        'TYPE': pick('Type','DESCRIPTION','TYPE'),
        'MOUNTING': pick('Mounting','MOUNT'),
        'ACCESSORIES1': pick('Accessories','ACCESSORIES','ACCESSORIES1'),
        'ACCESSORIES2': pick('Description','DESCRIPTION'),
        'REMARK': pick('Remark','Remarks','Notes','REMARK')
    }

def pad_numeric_only(tag):
    s = str(tag).strip()
    return s.zfill(2) if re.fullmatch(r'\d+', s) else s

_token_re = re.compile(r'(\d+|\D+)')
def tag_sort_key(tag):
    t = str(tag).strip()
    if t == '':
        return (3, '', ())
    if re.match(r'^[A-Za-z]', t):
        parts = _token_re.findall(t)
        seq = tuple((0, p.lower()) if not p.isdigit() else (1, int(p)) for p in parts)
        return (0, t.lower(), seq)
    m = re.match(r'^(\d+)', t)
    if m:
        num = int(m.group(1))
        rest = t[m.end():].lower()
        return (1, num, rest)
    return (2, t.lower(), ())

def neck_size_key(val):
    m = re.search(r'(\d+(\.\d+)?)', str(val))
    return float(m.group(1)) if m else float('inf')

def mounting_rank(val):
    if not isinstance(val, str):
        return 2
    v = val.lower()
    lay_variants = ['lay-in','lay in','in lay','in laying','in-lay','layin','laying','inlay','lay']
    for variant in lay_variants:
        if variant in v:
            return 0
    if 'surface' in v:
        return 1
    return 2

def to_float_safe(x):
    try:
        s = str(x).strip()
        return float(s) if s!='' else 0.0
    except:
        return 0.0

_tag_prefix_re = re.compile(r'^([A-Za-z]+)')
def _tag_prefix(tag):
    m = _tag_prefix_re.match(str(tag).strip())
    return m.group(1).upper() if m else str(tag).strip().upper()

def summarize_tag_prefixes_with_hash(tags_series):
    prefixes = []
    for t in tags_series:
        t = str(t).strip()
        if t == '':
            continue
        prefixes.append(_tag_prefix(t))
    if not prefixes:
        return '', ' (TOTAL)'
    parts = [f"{p}#" for p in sorted(set(prefixes))]
    compact = "/".join(parts)
    return compact, f"{compact} (TOTAL)"

# Robust upload reader (returns (df, error_message))
def read_uploaded_file(uploaded_file):
    if uploaded_file is None:
        return None, "No file uploaded."

    name = getattr(uploaded_file, "name", str(uploaded_file)).lower()
    # Excel first
    if name.endswith(('.xls', '.xlsx')):
        try:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, dtype=str)
            if df.shape[1] == 0:
                return None, "Excel file read OK but no columns were detected."
            df = df.fillna('')
            df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
            return df, None
        except Exception as e:
            return None, f"Failed to read Excel file: {e}"

    # Otherwise treat as CSV-like
    try:
        uploaded_file.seek(0)
        raw = uploaded_file.read()
        if isinstance(raw, bytes):
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                try:
                    text = raw.decode('latin1')
                except Exception:
                    return None, "Failed to decode CSV file bytes as UTF-8 or Latin-1."
        else:
            text = str(raw)

        lines = [ln for ln in text.splitlines() if ln.strip() != '']
        if len(lines) == 0:
            return None, "CSV appears to be empty (no non-blank lines)."

        sample = "\n".join(lines[:10])

        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            delim = dialect.delimiter
        except Exception:
            for d in [',',';','\t','|']:
                if d in sample:
                    delim = d
                    break
            else:
                delim = ','

        try:
            df = pd.read_csv(io.StringIO(text), sep=delim, dtype=str, engine='python')
        except Exception as e:
            try:
                df = pd.read_csv(io.StringIO(text), sep=None, engine='python', dtype=str)
            except Exception as e2:
                return None, f"Failed to parse CSV: {e} (fallback error: {e2})"

        if df.shape[1] == 0:
            return None, "CSV read OK but no columns were detected."
        df = df.fillna('')
        df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
        return df, None

    except Exception as e:
        return None, f"Unexpected failure reading file: {e}"

# Styling & excel save (two sheets)
def style_and_save_two_sheets(source_df, takeoff_df, excel_bytes_io,
                         header_fill="#CFE2F3", subtotal_fill="#FFFF00", grand_fill="#CFE2F3",
                         bold_cols = ['PRODUCT','SCHEDULE MODEL','TAG','MODULE SIZE','TYPE','ACCESSORIES1']):
    wb = Workbook()
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # RawData sheet (first)
    ws_raw = wb.active
    ws_raw.title = "RawData"
    for c_idx, col in enumerate(source_df.columns, start=1):
        cell = ws_raw.cell(row=1, column=c_idx, value=col)
        cell.fill = PatternFill(start_color=header_fill.replace('#',''), end_color=header_fill.replace('#',''), fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    for r_idx, row in enumerate(source_df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_raw.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
    for i, col in enumerate(source_df.columns, 1):
        try:
            max_len = max(source_df[col].astype(str).map(len).max(), len(col)) + 2
        except Exception:
            max_len = len(col) + 2
        ws_raw.column_dimensions[get_column_letter(i)].width = max_len

    # TakeoffNSW sheet
    ws = wb.create_sheet(title="TakeoffNSW")
    for c_idx, col in enumerate(takeoff_df.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = PatternFill(start_color=header_fill.replace('#',''), end_color=header_fill.replace('#',''), fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for r_idx, row in enumerate(takeoff_df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            colname = takeoff_df.columns[c_idx-1]
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if colname in bold_cols and str(value).strip() != '':
                cell.font = Font(bold=True)
            if colname == 'QTY' and str(value).strip() != '':
                try:
                    tag_val = str(row[list(takeoff_df.columns).index('TAG')]).strip() if 'TAG' in takeoff_df.columns else ''
                except Exception:
                    tag_val = ''
                prod_val = str(row[0]).strip()
                if tag_val.endswith('(TOTAL)') or prod_val == 'Grand Total' or ('=' in prod_val and prod_val.split('=')[-1].strip().replace('.','',1).isdigit()):
                    cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=False)
            if colname == 'TAG' and isinstance(value, str) and value.endswith('(TOTAL)'):
                for name in ['PRODUCT','TAG','QTY']:
                    if name in takeoff_df.columns:
                        i = list(takeoff_df.columns).index(name) + 1
                        ccell = ws.cell(row=r_idx, column=i)
                        ccell.fill = PatternFill(start_color=subtotal_fill.replace('#',''), end_color=subtotal_fill.replace('#',''), fill_type="solid")
                        ccell.font = Font(bold=True)
                        ccell.border = border
            if value == 'Grand Total':
                for col_i in range(1, len(takeoff_df.columns)+1):
                    ccell = ws.cell(row=r_idx, column=col_i)
                    ccell.fill = PatternFill(start_color=grand_fill.replace('#',''), end_color=grand_fill.replace('#',''), fill_type="solid")
                    ccell.font = Font(bold=True)
                    ccell.border = border

    ws.freeze_panes = 'A2'
    for i, col in enumerate(takeoff_df.columns, 1):
        try:
            max_len = max(takeoff_df[col].astype(str).map(len).max(), len(col)) + 2
        except Exception:
            max_len = len(col) + 2
        ws.column_dimensions[get_column_letter(i)].width = max_len

    wb.save(excel_bytes_io)
    excel_bytes_io.seek(0)
    return excel_bytes_io

# ---------------------------
# Corrected build_takeoff (with manual tag normalization fix)
# ---------------------------

def build_takeoff(df_in,
                  mapping,
                  apply_brand_autofill=True,
                  keep_model_blank=True,
                  replace_empty_with_dot=False,
                  aggregate_duplicates=True,
                  product_order=None,
                  tag_order=None,
                  manual_tags_by_product=None):
    """
    Returns: src_df, final_df, integrity
    manual_tags_by_product: dict mapping normalized product -> list of tag strings (free-form). This function
    normalizes those saved tag strings (pad numeric-only to 2 digits and lowercase) to match df['TAG'] normalization.
    """
    target_cols = ['PRODUCT','SCHEDULE BRAND','SCHEDULE MODEL','BRAND','MODEL','QTY','TAG',
                   'NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2','REMARK']

    # Map columns -> this will be our source (RawData)
    src_df = pd.DataFrame()
    for c in target_cols:
        src = mapping.get(c)
        if src and src in df_in.columns:
            src_df[c] = df_in[src].astype(str).fillna('')
        else:
            src_df[c] = ''

    # Working copy
    df = src_df.copy()

    # Brand autofill
    if apply_brand_autofill:
        prod_upper = df['PRODUCT'].astype(str).str.upper()
        df['BRAND'] = df.get('BRAND','').astype(str)
        mask_adgrd = prod_upper.str.contains('AD-GRD', na=False) & (df['BRAND'].astype(str).str.strip()=='')
        mask_fan = prod_upper.str.contains('FAN', na=False) & (df['BRAND'].astype(str).str.strip()=='')
        mask_split = prod_upper.str.contains('SPLIT SYSTEM', na=False) & (df['BRAND'].astype(str).str.strip()=='')
        df.loc[mask_adgrd, 'BRAND'] = 'PRICE'
        df.loc[mask_fan, 'BRAND'] = 'LOREN COOK'
        df.loc[mask_split, 'BRAND'] = 'SAMSUNG'
    else:
        if mapping.get('BRAND') and mapping['BRAND'] in df_in.columns:
            df['BRAND'] = df_in[mapping['BRAND']].astype(str).fillna('')
        else:
            df['BRAND'] = ''

    # MODEL blank handling
    if keep_model_blank:
        df['MODEL'] = ''
    else:
        if mapping.get('MODEL') and mapping['MODEL'] in df_in.columns:
            df['MODEL'] = df_in[mapping['MODEL']].astype(str).fillna('')

    # Replace empties with dot if requested
    if replace_empty_with_dot:
        df = df.replace('', '.').replace('nan','.')

    # numeric qty and tag normalization
    df['QTY'] = df['QTY'].apply(to_float_safe)
    df['TAG'] = df['TAG'].astype(str).apply(lambda t: pad_numeric_only(t.strip()))

    # optional aggregation of exact duplicate attribute rows
    if aggregate_duplicates:
        agg_keys = ['PRODUCT','TAG','SCHEDULE BRAND','SCHEDULE MODEL','NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2']
        for k in agg_keys:
            if k not in df.columns:
                df[k] = ''
        grouped = df.groupby(agg_keys, dropna=False, as_index=False).agg({
            'BRAND': 'first',
            'MODEL': 'first',
            'QTY': 'sum',
            'REMARK': lambda x: '; '.join([v for v in x.astype(str).unique() if v and v!='nan'])
        })
        new_df = pd.DataFrame()
        for c in target_cols:
            new_df[c] = grouped.get(c, '')
        df = new_df

    # ordering keys
    def _display_name_for_order(x, blank_placeholder):
        s = '' if (pd.isna(x) or str(x).strip() == '') else str(x).strip()
        return blank_placeholder if s == '' else s

    if product_order:
        prod_norm_order = [_display_name_for_order(p, '<<BLANK PRODUCT>>') for p in product_order]
        prod_rank = {p: i for i, p in enumerate(prod_norm_order)}
        df['_prod_key'] = df['PRODUCT'].apply(lambda x: prod_rank.get(_display_name_for_order(x, '<<BLANK PRODUCT>>'), len(prod_rank)))
    else:
        df['_prod_key'] = df['PRODUCT'].fillna('').astype(str).str.lower()

    # normalize manual_tags_by_product into manual_by_product using same normalization as df['TAG']
    manual_by_product = {}
    if manual_tags_by_product and isinstance(manual_tags_by_product, dict):
        for k, v in manual_tags_by_product.items():
            if not isinstance(v, list):
                continue
            pnorm = str(k).strip().lower()
            normalized_tags = []
            for x in v:
                if x is None:
                    continue
                xstr = str(x).strip()
                xnorm = pad_numeric_only(xstr)
                xnorm = xnorm.lower()
                normalized_tags.append(xnorm)
            manual_by_product[pnorm] = normalized_tags

    global_tag_rank = {}
    if tag_order and isinstance(tag_order, list):
        normalized = [str(t).strip() for t in tag_order]
        global_tag_rank = {t: i for i, t in enumerate(normalized)}

    def compute_tag_key(row):
        prod = str(row.get('PRODUCT','')).strip().lower()
        raw_tag = str(row.get('TAG','')).strip()
        tag_norm = pad_numeric_only(raw_tag).lower()
        L = manual_by_product.get(prod)
        if L:
            if tag_norm in L:
                return ('manual', L.index(tag_norm))
            else:
                return ('manual_auto', tag_sort_key(raw_tag))
        if raw_tag in global_tag_rank:
            return ('global', global_tag_rank[raw_tag])
        return ('auto', tag_sort_key(raw_tag))

    df['_tag_key'] = df.apply(lambda r: compute_tag_key(r), axis=1)
    df['_neck'] = df['NECK SIZE'].apply(neck_size_key)
    df['_mount'] = df['MOUNTING'].apply(mounting_rank)

    # final sort
    df = df.sort_values(by=['_prod_key','_tag_key','_neck','_mount'], kind='mergesort').reset_index(drop=True)

    # build output rows
    out_rows = []
    grand_total = 0.0
    input_tags = set(df['TAG'].unique())

    for prod_key, group in df.groupby('_prod_key', sort=True):
        group = group.copy().reset_index(drop=True)
        if group.empty:
            continue
        prod_name = group.loc[0,'PRODUCT']
        is_adgrd = isinstance(prod_name, str) and 'AD-GRD' in prod_name.upper()
        if is_adgrd:
            first_of_product = True
            tags = group['TAG'].unique().tolist()
            for tag in tags:
                tag_group = group[group['TAG'] == tag]
                subtotal = tag_group['QTY'].sum()
                grand_total += subtotal
                for idx, r in tag_group.iterrows():
                    row = {c: r[c] if c in r.index else '' for c in target_cols}
                    row['PRODUCT'] = prod_name if first_of_product else ''
                    out_rows.append(row)
                    first_of_product = False
                compact, tagcol = summarize_tag_prefixes_with_hash(tag_group['TAG'])
                num_display = int(subtotal) if subtotal == int(subtotal) else float(f"{subtotal:.2f}")
                subtotal_row = {c: '' for c in target_cols}
                subtotal_row['PRODUCT'] = (f"{num_display} = {compact}" if compact else f"{tag}={num_display}")
                subtotal_row['TAG'] = (f"{compact} (TOTAL)" if compact else f"{tag} (TOTAL)")
                subtotal_row['QTY'] = subtotal
                out_rows.append(subtotal_row)
            out_rows.append({**{c:'' for c in target_cols}})
        else:
            first_of_product = True
            prod_total = 0.0
            for idx, r in group.iterrows():
                row = {c: r[c] if c in r.index else '' for c in target_cols}
                row['PRODUCT'] = prod_name if first_of_product else ''
                out_rows.append(row)
                first_of_product = False
                prod_total += r['QTY']
                grand_total += r['QTY']
            compact, tagcol = summarize_tag_prefixes_with_hash(group['TAG'])
            num_display = int(prod_total) if prod_total == int(prod_total) else float(f"{prod_total:.2f}")
            subtotal_row = {c: '' for c in target_cols}
            if compact:
                subtotal_row['PRODUCT'] = f"{num_display} = {compact}"
                subtotal_row['TAG'] = f"{compact} (TOTAL)"
            else:
                subtotal_row['PRODUCT'] = f"{prod_name}={int(prod_total) if prod_total==int(prod_total) else '{:.2f}'.format(prod_total)}"
                subtotal_row['TAG'] = f"{prod_name.split('=')[0] if '=' in prod_name else prod_name.split('=')[0]} (TOTAL)"
            subtotal_row['QTY'] = prod_total
            out_rows.append(subtotal_row)
            out_rows.append({**{c:'' for c in target_cols}})

    grand_row = {c: '' for c in target_cols}
    grand_row['PRODUCT'] = 'Grand Total'
    grand_row['TAG'] = 'Grand Total'
    grand_row['QTY'] = grand_total
    out_rows.append(grand_row)

    final_df = pd.DataFrame(out_rows, columns=target_cols)

    # integrity
    output_tags = set([t for t in final_df['TAG'].dropna() if isinstance(t,str) and t.strip()!='' and not t.endswith('(TOTAL)')])
    missing_tags = sorted(list(input_tags - output_tags))
    integrity = {
        "input_unique_tags": len(input_tags),
        "output_unique_tags": len([t for t in final_df['TAG'].unique() if t and not str(t).endswith('(TOTAL)')]),
        "all_tags_preserved": len(missing_tags) == 0,
        "missing_tags_sample": missing_tags[:20]
    }

    return src_df, final_df, integrity

# ---------------------------
# Small helpers used by Triune UI (normalize product text, ordering)
# ---------------------------

def normalize_product_text(s: str) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    return t.lower()

def _norm_key(s: str) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()

def safe_rerun():
    try:
        rerun_fn = getattr(st, "experimental_rerun", None)
        if callable(rerun_fn):
            rerun_fn()
            return
        try:
            from streamlit.runtime.scriptrunner import RerunException
            raise RerunException()
        except Exception:
            return
    except Exception:
        return

# ---------------------------
# Manual-only Editors (textarea + Apply)
# ---------------------------

def render_manual_product_editor(preview_df: pd.DataFrame):
    product_values = preview_df["PRODUCT"].fillna("").astype(str).tolist()
    seen = set()
    products = []
    for p in product_values:
        if p not in seen:
            seen.add(p)
            products.append(p)

    if not products:
        st.info("No products found to order.")
        return

    with st.expander("Manual PRODUCT ordering — one product per line (click Apply to save)", expanded=False):
        fallback_key = "manual_order_textarea"
        default_text = "\n".join(products)
        pref = st.session_state.get('triune_manual_products_original')
        if pref and isinstance(pref, list) and len(pref) > 0:
            default_text = "\n".join(pref)
        pasted = st.text_area("Manual product order (one product per line):", value=st.session_state.get(fallback_key, default_text), height=240, key=fallback_key)
        lines = [ln.strip() for ln in pasted.splitlines() if ln.strip() != ""]
        preview_order = lines or products
        st.markdown("**Preview (unsaved) product order:**")
        st.write(preview_order)

        manual_norm = []
        manual_orig = []
        seen2 = set()
        for orig in preview_order:
            pnorm = normalize_product_text(orig)
            if pnorm != "" and pnorm not in seen2:
                seen2.add(pnorm)
                manual_norm.append(pnorm)
                manual_orig.append(orig)

        col1, col2 = st.columns([1,1])
        if col1.button("Apply product order"):
            if manual_norm:
                st.session_state['triune_manual_products'] = manual_norm
                st.session_state['triune_manual_products_original'] = manual_orig
                st.success(f"Product order applied — {len(manual_norm)} products saved.")
                safe_rerun()
            else:
                st.session_state.pop('triune_manual_products', None)
                st.session_state.pop('triune_manual_products_original', None)
                st.info("Cleared manual product order (nothing saved).")
                safe_rerun()

        if col2.button("Reset manual product order"):
            st.session_state.pop('triune_manual_products', None)
            st.session_state.pop('triune_manual_products_original', None)
            st.success("Manual product order cleared.")
            safe_rerun()

def render_manual_per_product_tag_editor(preview_df: pd.DataFrame):
    detected_products = preview_df["PRODUCT"].fillna("").astype(str).tolist()
    seen = set(); prod_list = []
    for p in detected_products:
        pn = _norm_key(p)
        if pn not in seen:
            seen.add(pn); prod_list.append(p)

    manual_prod_norm = st.session_state.get('triune_manual_products') or []
    manual_prod_display = []
    if manual_prod_norm:
        orig_map = {}
        for orig in prod_list:
            orig_map[_norm_key(orig)] = orig
        for mp in manual_prod_norm:
            if mp in orig_map:
                manual_prod_display.append(orig_map[mp])
        for p in prod_list:
            if p not in manual_prod_display:
                manual_prod_display.append(p)
        products_for_ui = manual_prod_display
    else:
        products_for_ui = prod_list

    if not products_for_ui:
        st.info("No products found for per-product tag editor.")
        return

    prev = st.session_state.get('triune_perprod_selected_product')
    sel_index = 0
    if prev and prev in products_for_ui:
        sel_index = products_for_ui.index(prev)
    sel_prod = st.selectbox("Select product to edit tags for", options=products_for_ui, index=sel_index, key="perprod_tag_select")
    st.session_state['triune_perprod_selected_product'] = sel_prod
    sel_prod_norm = _norm_key(sel_prod)

    # collect tags for selected product (unique order-preserving)
    tags_raw = preview_df.loc[preview_df["PRODUCT"].fillna("").astype(str) == sel_prod, "TAG"].fillna("").astype(str).tolist()
    seen_t = set(); tags = []
    for t in tags_raw:
        if t not in seen_t:
            seen_t.add(t); tags.append(t)

    if not tags:
        st.info(f"No TAGs found for product: {sel_prod}")
        return

    with st.expander(f"Edit TAG order for product: {sel_prod} (click Apply to save)", expanded=False):
        fallback_key = f"manual_tag_order_{_norm_key(sel_prod)}"
        default_text = "\n".join(tags)
        saved_orig_map = st.session_state.get('triune_manual_tags_by_product_orig') or {}
        if sel_prod_norm in saved_orig_map:
            default_text = "\n".join(saved_orig_map[sel_prod_norm])
        pasted = st.text_area("Manual tag order for this product (one tag per line):", value=st.session_state.get(fallback_key, default_text), height=200, key=fallback_key)
        lines = [ln.strip() for ln in pasted.splitlines() if ln.strip() != ""]
        preview_order = lines or tags
        st.markdown("**Preview (unsaved) tag order for product:**")
        st.write(preview_order)

        manual_norm_list = []
        manual_orig_list = []
        for tv in preview_order:
            tn = _norm_key(tv)
            if tn != "":
                manual_norm_list.append(tn)
                manual_orig_list.append(tv)

        col1, col2 = st.columns([1,1])
        if col1.button(f"Apply TAG order for product: {sel_prod}"):
            if 'triune_manual_tags_by_product' not in st.session_state:
                st.session_state['triune_manual_tags_by_product'] = {}
            if 'triune_manual_tags_by_product_orig' not in st.session_state:
                st.session_state['triune_manual_tags_by_product_orig'] = {}
            if manual_norm_list:
                st.session_state['triune_manual_tags_by_product'][sel_prod_norm] = manual_norm_list
                st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm] = manual_orig_list
                st.success(f"Saved manual TAG order for product: {sel_prod} ({len(manual_norm_list)} tags).")
                safe_rerun()
            else:
                if sel_prod_norm in st.session_state.get('triune_manual_tags_by_product', {}):
                    del st.session_state['triune_manual_tags_by_product'][sel_prod_norm]
                if sel_prod_norm in st.session_state.get('triune_manual_tags_by_product_orig', {}):
                    del st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm]
                st.info(f"Cleared manual TAG order for product: {sel_prod} (nothing saved).")
                safe_rerun()

        if col2.button(f"Reset TAG order for product: {sel_prod}"):
            if 'triune_manual_tags_by_product' in st.session_state and sel_prod_norm in st.session_state['triune_manual_tags_by_product']:
                del st.session_state['triune_manual_tags_by_product'][sel_prod_norm]
            if 'triune_manual_tags_by_product_orig' in st.session_state and sel_prod_norm in st.session_state['triune_manual_tags_by_product_orig']:
                del st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm]
            st.success(f"Cleared saved TAG order for product: {sel_prod}.")
            safe_rerun()

# ---------------------------
# Triune Data Unit helper functions
# ---------------------------

def normalize_text(s: str) -> str:
    return re.sub(r"[\s_\-\.]+", " ", str(s).strip().lower())

def find_column(raw_cols: List[str], *aliases: str) -> Optional[str]:
    raw_norm = {c: normalize_text(c) for c in raw_cols}
    for alias in aliases:
        a_norm = normalize_text(alias)
        for rc, rn in raw_norm.items():
            if rn == a_norm:
                return rc
    for alias in aliases:
        a_norm = normalize_text(alias)
        for rc, rn in raw_norm.items():
            if a_norm in rn or rn in a_norm:
                return rc
    alias_tokens = set()
    for alias in aliases:
        alias_tokens.update(normalize_text(alias).split())
    best = None
    best_score = 0
    for rc, rn in raw_norm.items():
        tokens = set(rn.split())
        score = len(tokens & alias_tokens)
        if score > best_score:
            best_score = score
            best = rc
    if best_score > 0:
        return best
    return None

def load_file_to_df_simple_for_units(uploaded_file):
    if uploaded_file is None:
        return None
    df, err = read_uploaded_file(uploaded_file)
    if err:
        return None
    return df

def detect_unit_column(df: pd.DataFrame) -> Optional[str]:
    for c in df.columns:
        if re.search(r"(^unit$)|unit(s)?|apt|apartment|flat|room|suite|unit_id|unitno|unit_no", c, re.IGNORECASE):
            return c
    return None

def apply_raw_column_mapping(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = raw_df.copy()
    cols = list(df.columns)
    lowered = {c.lower(): c for c in cols}
    exact_rules = {
        "subject": "PRODUCT",
        "page index": "Page Index",
        "label": "TAG",
        "manufacturer": "BRAND",
        "face size": "MODULE SIZE",
        "description": "ACCESSORIES2",
        "accessories": "ACCESSORIES1",
        "accessories1": "ACCESSORIES1"
    }
    exact_map = {}
    for low_key, dest in exact_rules.items():
        if low_key in lowered:
            orig = lowered[low_key]
            if orig != dest:
                exact_map[orig] = dest
    if exact_map:
        df = df.rename(columns=exact_map)
        cols = list(df.columns)
    mapping_pairs = {
        "QTY": ("quantity", "qty", "count", "qty.", "q'ty"),
        "BRAND": ("brand", "make", "mfr"),
        "MODEL": ("model", "catalog", "cat no", "catalog no"),
        "TAG": ("tag", "tag id", "label", "ref", "mark"),
        "NECK SIZE": ("neck size", "neck"),
        "MODULE SIZE": ("module size", "face size", "module", "face"),
        "DUCT SIZE": ("duct size", "duct"),
        "CFM": ("cfm", "airflow"),
        "TYPE": ("type", "desc", "description"),
        "MOUNTING": ("mounting", "install", "mount"),
        "ACCESSORIES1": ("accessories", "accessories1", "accessory 1", "accessory"),
        "ACCESSORIES2": ("accessories2", "accessory 2", "description", "desc"),
        "REMARK": ("remark", "remarks", "note", "notes"),
        "UNITS": ("unit", "units", "zone", "area", "apt", "apartment")
    }
    rename_dict = {}
    cols = list(df.columns)
    for target_col, aliases in mapping_pairs.items():
        if target_col in df.columns:
            continue
        found = find_column(cols, *aliases)
        if found and found not in rename_dict and found != target_col:
            rename_dict[found] = target_col
    if rename_dict:
        df = df.rename(columns=rename_dict)
    return df

def clean_unit_matrix(df: pd.DataFrame, unit_col_hint: Optional[str] = None) -> (pd.DataFrame, str):
    df2 = df.copy()
    unit_col = unit_col_hint or detect_unit_column(df2) or df2.columns[0]
    df2[unit_col] = df2[unit_col].fillna("").astype(str).str.strip()
    df2 = df2[~df2[unit_col].str.upper().isin(["TOTAL", "GRAND TOTAL", "SUMMARY", "ALL"])]
    return df2.reset_index(drop=True), unit_col

def guess_multiplier_column(unit_df: pd.DataFrame, unit_col: str) -> Optional[str]:
    candidates = [c for c in unit_df.columns if c != unit_col]
    for c in candidates:
        sample = unit_df[c].dropna().astype(str).str.replace(",", "").str.strip()
        if sample.size and any(s.replace(".", "", 1).isdigit() for s in sample[:50]):
            return c
    return candidates[0] if candidates else None

def build_data_unit_sheet(
    raw_df: pd.DataFrame,
    unit_df: pd.DataFrame,
    raw_unit_col: str,
    matrix_unit_col: str,
    multiplier_col: Optional[str],
    selected_units: Optional[list] = None,
    include_empty: bool = True,
    default_multiplier: int = 1
) -> pd.DataFrame:
    raw = raw_df.copy()
    mat = unit_df.copy()
    raw[raw_unit_col] = raw[raw_unit_col].fillna("").astype(str).str.strip()
    mat[matrix_unit_col] = mat[matrix_unit_col].fillna("").astype(str).str.strip()
    multiplier_map = {}
    if multiplier_col:
        for _, r in mat.iterrows():
            u = str(r.get(matrix_unit_col, "")).strip()
            v = r.get(multiplier_col, "")
            try:
                if pd.isna(v) or str(v).strip() == "":
                    continue
                num = int(float(str(v).replace(",", "").strip()))
                multiplier_map[u] = num
            except Exception:
                continue
    org_count_col = None
    for name in ["Org Count", "OrgCount", "ORIGINAL COUNT", "Count", "QTY", "Qty", "qty", "QTY/UNIT"]:
        if name in raw.columns:
            org_count_col = name
            break
    if org_count_col is None:
        for c in raw.columns:
            sample = raw[c].dropna().astype(str).str.replace(",", "").str.strip()
            if sample.size and all(s.replace(".", "", 1).isdigit() for s in sample[:50]):
                org_count_col = c
                break
    if org_count_col is None:
        raw["Org Count"] = 1
    else:
        raw["Org Count"] = pd.to_numeric(raw[org_count_col].fillna("0").astype(str).str.replace(",", "").str.strip(), errors="coerce").fillna(0).astype(int)
    def get_mult(u):
        if u == "" or pd.isna(u):
            return default_multiplier
        return multiplier_map.get(u, default_multiplier)
    raw["__unit_multiplier__"] = raw[raw_unit_col].apply(get_mult).astype(int)
    raw["Count"] = raw["Org Count"].astype(int) * raw["__unit_multiplier__"]
    if selected_units:
        sel = set(selected_units)
        raw_units_vals = raw[raw_unit_col].fillna("").astype(str).str.strip()
        mask = raw_units_vals.isin(sel)
        if "<<EMPTY UNIT>>" in sel:
            mask = mask | (raw_units_vals == "")
        raw = raw[mask].copy()
    out_unit_name = "UNITS"
    raw[out_unit_name] = raw[raw_unit_col].replace({"": "<<EMPTY UNIT>>"})
    required_order = [
        "PRODUCT", "Page Index", "TAG", "Org Count", "Count", "BRAND", "MODEL",
        "NECK SIZE", "MODULE SIZE", "DUCT SIZE", "CFM", "TYPE", "MOUNTING",
        "ACCESSORIES1", "ACCESSORIES2", "REMARK", out_unit_name, "DAMPER TYPE"
    ]
    out = raw.copy()
    for col in required_order:
        if col not in out.columns:
            out[col] = ""
    other_cols = [c for c in out.columns if c not in required_order and c != "__unit_multiplier__"]
    final_cols = required_order + other_cols
    if "__unit_multiplier__" in out.columns:
        out = out.drop(columns=["__unit_multiplier__"])
    return out[final_cols].copy()

def bytes_from_df_excel(df: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data Unit Sheet")
    return out.getvalue()

def create_zip_bytes_from_map(dfs_map: dict, export_fn) -> bytes:
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for unit, df in dfs_map.items():
            safe = "".join(c if c.isalnum() or c in " -_." else "_" for c in str(unit))[:120] or "unit"
            fname = f"{safe}.xlsx"
            zf.writestr(fname, export_fn(df))
    return mem.getvalue()

def export_styled_excel_bytes(df: pd.DataFrame) -> bytes:
    return bytes_from_df_excel(df)

# ---------------------------
# Streamlit UI: Two tabs — Takeoff & Data Units
# ---------------------------

APP_TITLE = "TakeoffNSW + Data Unit Tools (Manual ordering)"

st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)
st.markdown("Upload raw takeoff and convert to TakeoffNSW format (Triune-compatible), or build Data Unit Sheets (Triune helpers). Manual PRODUCT ordering and per-product TAG ordering included.")

tabs = st.tabs(["Takeoff", "Data Units"])

# ---------- TAB: Takeoff ----------
with tabs[0]:
    st.header("TakeoffNSW Formatter (Triune-compatible output)")
    uploaded = st.file_uploader("Upload CSV or Excel (raw takeoff)", type=['csv','xlsx','xls'], key="nsw_upload")

    st.sidebar.header("Options")
    apply_brand = st.sidebar.checkbox("Apply brand autofill rules (AD-GRD/FAN/SPLIT SYSTEM)", value=True)
    keep_model_blank = st.sidebar.checkbox("Keep MODEL blank", value=True)
    replace_empty = st.sidebar.checkbox("Replace empty cells with '.'", value=False)
    aggregate_duplicates = st.sidebar.checkbox("Aggregate exact duplicate attribute rows (sum QTY)", value=True)

    st.sidebar.markdown("### Colors")
    header_color = st.sidebar.color_picker("Header fill", "#CFE2F3")
    subtotal_color = st.sidebar.color_picker("Subtotal fill", "#FFFF00")
    grand_color = st.sidebar.color_picker("Grand total fill", "#CFE2F3")

    st.sidebar.markdown("### Bold columns (comma-separated)")
    bold_input = st.sidebar.text_input("Columns to bold (exact names)", value="PRODUCT,SCHEDULE MODEL,TAG,MODULE SIZE,TYPE,ACCESSORIES1")
    bold_cols = [c.strip() for c in bold_input.split(',') if c.strip()]

    st.sidebar.markdown("### Mapping JSON (optional)")
    mapping_json = st.sidebar.text_area("Mapping JSON (target -> source). Leave empty to auto-detect.", value="", height=140, key="nsw_mapjson")
    if st.sidebar.button("Load mapping from JSON", key="nsw_loadmapbtn"):
        try:
            m = json.loads(mapping_json)
            st.session_state['nsw_mapping'] = m
            st.sidebar.success("Mapping loaded into session.")
        except Exception as e:
            st.sidebar.error(f"Invalid JSON: {e}")

    if uploaded:
        df_in, read_err = read_uploaded_file(uploaded)
        if read_err:
            st.error(f"Failed to read file: {read_err}")
        else:
            auto_map = detect_columns(df_in)
            mapping_use = auto_map.copy()
            if st.session_state.get('nsw_mapping'):
                for k,v in st.session_state['nsw_mapping'].items():
                    mapping_use[k] = v

            st.subheader("Column mapping (target → source)")
            with st.form("mapping_form_nsw"):
                cols = ['PRODUCT','SCHEDULE BRAND','SCHEDULE MODEL','BRAND','MODEL','QTY','TAG','NECK SIZE','MODULE SIZE','DUCT SIZE','TYPE','MOUNTING','ACCESSORIES1','ACCESSORIES2','REMARK']
                user_map = {}
                for c in cols:
                    choices = [''] + list(df_in.columns)
                    default = mapping_use.get(c) if mapping_use.get(c) in df_in.columns else ''
                    user_map[c] = st.selectbox(c, choices, index=(choices.index(default) if default in choices else 0), key=f"map_{c}_nsw")
                save_map = st.form_submit_button("Apply mapping")
            if save_map:
                st.session_state['nsw_mapping'] = user_map
                st.success("Mapping applied and saved.")

            mapping_use_final = st.session_state.get('nsw_mapping') or mapping_use

            # Preview mapped source
            preview_build = pd.DataFrame()
            for c in cols:
                src = (mapping_use_final.get(c) if mapping_use_final.get(c) else auto_map.get(c))
                if src and src in df_in.columns:
                    preview_build[c] = df_in[src].astype(str).fillna('')
                else:
                    preview_build[c] = ''

            st.markdown("**Preview of mapped data (first 20 rows)**")
            st.dataframe(preview_build.head(20), height=300)

            # Manual product & per-product tag editors
            st.markdown("### Manual PRODUCT ordering")
            st.caption("One product per line. Click Apply to save a manual product order which will be used when generating the takeoff.")
            render_manual_product_editor(preview_build)

            st.markdown("### Per-product TAG ordering")
            st.caption("Select a product, then paste one tag per line in the textarea and click Apply. This per-product ordering overrides global tag sort for that product.")
            render_manual_per_product_tag_editor(preview_build)

            st.markdown("---")
            st.markdown("This app uses automatic product/tag ordering **unless** you supply a manual product list (Apply) or per-product tag lists (Apply).")

            if st.button("Generate Takeoff files (preview + download)", key="gen_takeoff"):
                with st.spinner("Building takeoff..."):
                    prod_ord = st.session_state.get('triune_manual_products_original') if st.session_state.get('triune_manual_products_original') else None
                    manual_tags = st.session_state.get('triune_manual_tags_by_product') or {}

                    src_df, final_df, integrity = build_takeoff(
                        df_in,
                        mapping_use_final,
                        apply_brand_autofill=apply_brand,
                        keep_model_blank=keep_model_blank,
                        replace_empty_with_dot=replace_empty,
                        aggregate_duplicates=aggregate_duplicates,
                        product_order=prod_ord,
                        tag_order=None,
                        manual_tags_by_product=manual_tags
                    )

                    st.subheader("Integrity check")
                    st.json(integrity)

                    st.subheader("Final Takeoff preview (first 200 rows)")
                    st.dataframe(final_df.head(200), height=400)

                    excel_io = io.BytesIO()
                    style_and_save_two_sheets(src_df, final_df, excel_io,
                                         header_fill=header_color, subtotal_fill=subtotal_color, grand_fill=grand_color,
                                         bold_cols=bold_cols)

                    csv_io = io.BytesIO()
                    csv_io.write(final_df.to_csv(index=False).encode('utf-8'))
                    csv_io.seek(0)

                    st.download_button("📥 Download Triune Takeoff Excel Output (RawData + Takeoff)", data=excel_io,
                                       file_name=f"Triune_Takeoff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    st.download_button("📥 Download CSV (Takeoff only)", data=csv_io,
                                       file_name=f"Triune_Takeoff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

                    st.success("Files generated. Download above.")
    else:
        st.info("Upload a CSV or Excel file to begin the Takeoff conversion.")

# ---------- TAB: Data Units ----------
with tabs[1]:
    st.header("Data Unit Sheet Builder (Triune helpers)")
    col1, col2 = st.columns(2)
    with col1:
        raw_file = st.file_uploader("Upload Raw Takeoff (CSV or Excel)", type=["csv", "xlsx", "xls"], key="du_raw")
    with col2:
        unit_file = st.file_uploader("Upload Unit Matrix (CSV or Excel)", type=["csv", "xlsx", "xls"], key="du_matrix")

    if not raw_file or not unit_file:
        st.info("Upload both Raw Takeoff and Unit Matrix files to continue.")
    else:
        raw_df = load_file_to_df_simple_for_units(raw_file)
        unit_df = load_file_to_df_simple_for_units(unit_file)
        if raw_df is None or unit_df is None:
            st.error("Couldn't read one of the uploaded files. Please check file formats/encoding.")
        else:
            raw_df = apply_raw_column_mapping(raw_df)
            raw_unit_detected = detect_unit_column(raw_df)
            unit_matrix_clean, matrix_unit_detected = clean_unit_matrix(unit_df)

            st.markdown("### Detected columns")
            st.write(f"Raw unit column (auto-detected): **{raw_unit_detected or 'None'}**")
            st.write(f"Unit matrix unit column (auto-detected): **{matrix_unit_detected}**")

            raw_unit_col = st.selectbox("Select unit column in raw file", options=list(raw_df.columns),
                                        index=list(raw_df.columns).index(raw_unit_detected) if raw_unit_detected in raw_df.columns else 0,
                                        key="du_raw_unit_col")
            matrix_unit_col = st.selectbox("Select unit column in unit matrix", options=list(unit_matrix_clean.columns),
                                           index=list(unit_matrix_clean.columns).index(matrix_unit_detected) if matrix_unit_detected in unit_matrix_clean.columns else 0,
                                           key="du_matrix_unit_col")

            guessed_mult = guess_multiplier_column(unit_matrix_clean, matrix_unit_col)
            mult_options = [None] + list(unit_matrix_clean.columns)
            mult_index = 1 + list(unit_matrix_clean.columns).index(guessed_mult) if guessed_mult in unit_matrix_clean.columns else 0
            multiplier_col = st.selectbox("Select multiplier column in unit matrix (units per type)", options=mult_options,
                                         index=mult_index, key="du_multcol")
            if multiplier_col is None:
                st.warning("No multiplier column selected — multipliers default to 1 when unmatched.")

            units_series = unit_matrix_clean[matrix_unit_col].dropna().astype(str).str.strip()
            raw_units = raw_df[raw_unit_col].fillna("").astype(str).str.strip().unique().tolist()
            has_empty = any(u == "" for u in raw_df[raw_unit_col].fillna("").astype(str).str.strip().tolist())
            units_list = sorted(set(units_series.tolist() + [u for u in raw_units if u != ""]))
            if has_empty:
                units_list = ["<<EMPTY UNIT>>"] + units_list

            st.markdown("### Choose units to include (checkboxes)")
            select_all = st.checkbox("Select all units", value=True, key="du_selall")
            cols_chk = st.columns(3)
            selected_units = []
            for i, unit in enumerate(units_list):
                col = cols_chk[i % 3]
                default_val = True if select_all else False
                checked = col.checkbox(str(unit), value=default_val, key=f"du_unit_chk_{i}")
                if checked:
                    selected_units.append(unit)

            include_empty = st.checkbox("Include empty-unit rows (as <<EMPTY UNIT>>)", value=True, key="du_inc_empty")
            split_by_unit = st.checkbox("Split into separate Excel per unit (download as ZIP)", value=False, key="du_splitzip")

            st.markdown("### Output filename")
            default_name = "Data_Unit_Sheet"
            if split_by_unit:
                default_name = "takeoff_by_unit"
            file_name_input = st.text_input("Enter desired download filename (without extension):", value=default_name, key="du_outfilebase")
            def sanitize_filename_local(s: str) -> str:
                return "".join(c for c in s if c.isalnum() or c in " -_").strip() or default_name
            out_file_base = sanitize_filename_local(file_name_input)

            if st.button("Generate Data Unit Sheet", key="gen_du"):
                with st.spinner("Building Data Unit Sheet..."):
                    sel_units = selected_units if selected_units else None
                    final_df = build_data_unit_sheet(
                        raw_df=raw_df,
                        unit_df=unit_matrix_clean,
                        raw_unit_col=raw_unit_col,
                        matrix_unit_col=matrix_unit_col,
                        multiplier_col=multiplier_col,
                        selected_units=sel_units,
                        include_empty=include_empty,
                        default_multiplier=1
                    )
                    st.success("Data Unit Sheet created.")
                    st.write("Preview (first 200 rows):")
                    st.dataframe(final_df.head(200), use_container_width=True)

                    if split_by_unit:
                        groups = {u: g for u, g in final_df.groupby("UNITS")}
                        st.write("Files to be included in ZIP:")
                        summary = [{"unit": k, "rows": len(v)} for k, v in groups.items()]
                        st.table(pd.DataFrame(summary).sort_values("rows", ascending=False))
                        zip_bytes = create_zip_bytes_from_map(groups, export_styled_excel_bytes)
                        download_name = f"{out_file_base}.zip"
                        st.download_button("Download ZIP (per-unit Excels)", data=zip_bytes, file_name=download_name, mime="application/zip")
                    else:
                        excel_bytes = export_styled_excel_bytes(final_df)
                        download_name = f"{out_file_base}.xlsx"
                        st.download_button("Download combined Excel", data=excel_bytes, file_name=download_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
